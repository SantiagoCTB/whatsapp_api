from flask import Flask, request, jsonify, render_template, session, redirect, url_for, flash
import requests
import sqlite3
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv
import hashlib

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

META_TOKEN = os.getenv('META_TOKEN')
PHONE_NUMBER_ID = os.getenv('PHONE_NUMBER_ID')
VERIFY_TOKEN = os.getenv('VERIFY_TOKEN')

DB_PATH = 'database.db'
SESSION_TIMEOUT = 600  # 10 minutos
user_last_activity = {}
user_steps = {}

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Tabla mensajes
    c.execute('''
        CREATE TABLE IF NOT EXISTS mensajes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT,
            mensaje TEXT,
            tipo TEXT,
            timestamp TEXT
        )
    ''')
    
    # Tabla mensajes procesados
    c.execute('''
        CREATE TABLE IF NOT EXISTS mensajes_procesados (
            mensaje_id TEXT PRIMARY KEY
        )
    ''')

    # Tabla de usuarios
    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            rol TEXT NOT NULL
        )
    ''')

    # Tabla de reglas de automatización (con soporte para tipo y opciones interactivas)
    c.execute('''
        CREATE TABLE IF NOT EXISTS reglas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            step TEXT NOT NULL,
            input_text TEXT NOT NULL,
            respuesta TEXT NOT NULL,
            siguiente_step TEXT,
            tipo TEXT DEFAULT 'texto',
            opciones TEXT
        )
    ''')

    # Asegurar que la columna 'opciones' existe
    try:
        c.execute("ALTER TABLE reglas ADD COLUMN opciones TEXT")
    except:
        pass

    # Tabla de botones
    c.execute('''
        CREATE TABLE IF NOT EXISTS botones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mensaje TEXT NOT NULL
        )
    ''')

    # Crear usuario admin si no existe
    c.execute("SELECT * FROM usuarios WHERE username = 'admin'")
    if not c.fetchone():
        import hashlib
        password = 'admin123'
        hashed = hashlib.sha256(password.encode()).hexdigest()
        c.execute("INSERT INTO usuarios (username, password, rol) VALUES (?, ?, ?)",
                  ('admin', hashed, 'admin'))

    conn.commit()
    conn.close()

init_db()

def enviar_mensaje(numero, mensaje, tipo='bot', tipo_respuesta='texto', opciones=None):
    url = f"https://graph.facebook.com/v23.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {META_TOKEN}",
        "Content-Type": "application/json"
    }

    if tipo_respuesta == 'texto':
        data = {
            "messaging_product": "whatsapp",
            "to": numero,
            "type": "text",
            "text": {"body": mensaje}
        }

    elif tipo_respuesta == 'boton':
        # Hasta 3 botones
        buttons = [{"type": "reply", "reply": {"id": f"btn_{i}", "title": op}} for i, op in enumerate(opciones[:3])]
        data = {
            "messaging_product": "whatsapp",
            "to": numero,
            "type": "interactive",
            "interactive": {
                "type": "button",
                "body": {"text": mensaje},
                "action": {"buttons": buttons}
            }
        }

    elif tipo_respuesta == 'lista':
        sections = [{
            "title": "Opciones disponibles",
            "rows": [{"id": f"opt_{i}", "title": op, "description": ""} for i, op in enumerate(opciones)]
        }]
        data = {
            "messaging_product": "whatsapp",
            "to": numero,
            "type": "interactive",
            "interactive": {
                "type": "list",
                "header": {"type": "text", "text": "Menú"},
                "body": {"text": mensaje},
                "footer": {"text": "Selecciona una opción"},
                "action": {
                    "button": "Ver opciones",
                    "sections": sections
                }
            }
        }

    requests.post(url, headers=headers, json=data)
    guardar_mensaje(numero, mensaje, tipo)



def enviar_mensaje_boton(numero, mensaje, botones):
    url = f"https://graph.facebook.com/v23.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {META_TOKEN}",
        "Content-Type": "application/json"
    }

    botones_payload = [
        {"type": "reply", "reply": {"id": f"opcion_{i+1}", "title": btn.strip()}}
        for i, btn in enumerate(botones[:3])
    ]

    data = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "interactive",
        "interactive": {
            "type": "button",
            "body": {"text": mensaje},
            "action": {
                "buttons": botones_payload
            }
        }
    }

    requests.post(url, headers=headers, json=data)
    guardar_mensaje(numero, "[BOTÓN] " + mensaje, "bot")

def enviar_mensaje_lista(numero, mensaje, opciones):
    url = f"https://graph.facebook.com/v23.0/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {META_TOKEN}",
        "Content-Type": "application/json"
    }

    lista_payload = [{
        "title": "Opciones",
        "rows": [
            {"id": f"opcion_{i+1}", "title": opt.strip()}
            for i, opt in enumerate(opciones[:10])
        ]
    }]

    data = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "interactive",
        "interactive": {
            "type": "list",
            "body": {"text": mensaje},
            "action": {
                "button": "Seleccionar",
                "sections": lista_payload
            }
        }
    }

    requests.post(url, headers=headers, json=data)
    guardar_mensaje(numero, "[LISTA] " + mensaje, "bot")



def guardar_mensaje(numero, mensaje, tipo):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO mensajes (numero, mensaje, tipo, timestamp) VALUES (?, ?, ?, ?)",
              (numero, mensaje, tipo, str(datetime.now())))
    conn.commit()
    conn.close()

from flask import render_template, request, redirect, session, url_for

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    if "user" not in session or session["rol"] != "admin":
        return redirect(url_for("login"))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Cargar archivo Excel si se sube
    if request.method == 'POST':
        if 'archivo' in request.files:
            from openpyxl import load_workbook
            archivo = request.files['archivo']
            wb = load_workbook(archivo)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                step, input_text, respuesta, siguiente_step, tipo, opciones = fila

                c.execute("SELECT id FROM reglas WHERE step = ? AND input_text = ?", (step, input_text))
                existente = c.fetchone()
                if existente:
                    c.execute('''
                        UPDATE reglas
                        SET respuesta = ?, siguiente_step = ?, tipo = ?, opciones = ?
                        WHERE id = ?
                    ''', (respuesta, siguiente_step, tipo, opciones, existente[0]))
                else:
                    c.execute('''
                        INSERT INTO reglas (step, input_text, respuesta, siguiente_step, tipo, opciones)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (step, input_text, respuesta, siguiente_step, tipo, opciones))
            conn.commit()

        else:
            # Carga desde formulario manual
            step = request.form['step']
            input_text = request.form['input_text']
            respuesta = request.form['respuesta']
            siguiente_step = request.form['siguiente_step']
            tipo = request.form['tipo']
            opciones = request.form.get('opciones', None)

            c.execute("SELECT id FROM reglas WHERE step = ? AND input_text = ?", (step, input_text))
            existente = c.fetchone()
            if existente:
                c.execute('''
                    UPDATE reglas
                    SET respuesta = ?, siguiente_step = ?, tipo = ?, opciones = ?
                    WHERE id = ?
                ''', (respuesta, siguiente_step, tipo, opciones, existente[0]))
            else:
                c.execute('''
                    INSERT INTO reglas (step, input_text, respuesta, siguiente_step, tipo, opciones)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (step, input_text, respuesta, siguiente_step, tipo, opciones))

            conn.commit()

    c.execute("SELECT * FROM reglas ORDER BY step, id")
    reglas = c.fetchall()
    conn.close()

    return render_template('configuracion.html', reglas=reglas)

@app.route("/get_botones")
def get_botones():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, mensaje FROM botones ORDER BY id")
    botones = [{"id": row[0], "mensaje": row[1]} for row in c.fetchall()]
    conn.close()
    return jsonify(botones)

@app.route('/botones', methods=['GET', 'POST'])
def botones():
    if "user" not in session or session["rol"] != "admin":
        return redirect(url_for("login"))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if request.method == 'POST':
        if 'archivo' in request.files:
            from openpyxl import load_workbook
            archivo = request.files['archivo']
            wb = load_workbook(archivo)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                mensaje = fila[0]
                if mensaje:
                    c.execute("INSERT INTO botones (mensaje) VALUES (?)", (mensaje,))
            conn.commit()
        elif 'mensaje' in request.form:
            nuevo_mensaje = request.form['mensaje']
            c.execute("INSERT INTO botones (mensaje) VALUES (?)", (nuevo_mensaje,))
            conn.commit()

    c.execute("SELECT id, mensaje FROM botones ORDER BY id")
    botones = c.fetchall()
    conn.close()
    return render_template('botones.html', botones=botones)


@app.route('/eliminar_boton/<int:boton_id>', methods=['POST'])
def eliminar_boton(boton_id):
    if "user" not in session or session["rol"] != "admin":
        return redirect(url_for("login"))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM botones WHERE id = ?", (boton_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('botones'))

@app.route('/webhook', methods=['GET', 'POST'])
def webhook():
    if request.method == 'GET':
        if request.args.get('hub.verify_token') == VERIFY_TOKEN:
            return request.args.get('hub.challenge'), 200
        return 'Forbidden', 403

    data = request.get_json()

    if data.get('object'):
        for entry in data.get('entry', []):
            for change in entry.get('changes', []):
                value = change.get('value', {})
                messages = value.get('messages')
                if messages:
                    message = messages[0]
                    mensaje_id = message.get('id')
                    from_number = message['from']
                    text = message['text']['body'].strip().lower()

                    # Verificar duplicados
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute("SELECT 1 FROM mensajes_procesados WHERE mensaje_id = ?", (mensaje_id,))
                    if c.fetchone():
                        conn.close()
                        return jsonify({"status": "duplicate_ignored"})
                    c.execute("INSERT INTO mensajes_procesados (mensaje_id) VALUES (?)", (mensaje_id,))
                    conn.commit()
                    conn.close()

                    guardar_mensaje(from_number, text, 'cliente')

                    # Timeout
                    now = datetime.now()
                    last_time = user_last_activity.get(from_number)
                    if last_time and (now - last_time).total_seconds() > SESSION_TIMEOUT:
                        enviar_mensaje(from_number, "Muchas gracias por comunicarte con nosotros. La sesión se dará por terminada por inactividad. ¡Te esperamos nuevamente por aquí!")
                        user_steps.pop(from_number, None)
                    user_last_activity[from_number] = now

                    # Palabras clave para reiniciar
                    if text in ['reiniciar', 'volver al inicio', 'inicio', 'menú', 'menu', 'ayuda']:
                        user_steps.pop(from_number, None)
                        user_steps[from_number] = 'menu_principal'
                        enviar_mensaje(from_number, "Perfecto, volvamos a empezar.")

                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute("SELECT respuesta, siguiente_step FROM reglas WHERE step = 'menu_principal' AND input_text = 'iniciar'")
                        bienvenida = c.fetchone()
                        conn.close()

                        if bienvenida:
                            enviar_mensaje(from_number, bienvenida[0])
                            if bienvenida[1]:
                                user_steps[from_number] = bienvenida[1]
                        return jsonify({"status": "reiniciado"})

                    # Paso actual
                    step = user_steps.get(from_number)

                    # Si no hay paso, iniciar con mensaje de bienvenida
                    if not step:
                        step = 'menu_principal'
                        user_steps[from_number] = step

                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute("SELECT respuesta, siguiente_step FROM reglas WHERE step = ? AND input_text = ?", (step, 'iniciar'))
                        bienvenida = c.fetchone()
                        conn.close()

                        if bienvenida:
                            enviar_mensaje(from_number, bienvenida[0])
                            if bienvenida[1]:
                                user_steps[from_number] = bienvenida[1]  # ✅ se guarda el paso correcto
                        return jsonify({"status": "sent_welcome"})

                    # Lógica de cotización con medidas
                    try:
                        if step == 'barra_medida':
                            medida = int(text)
                            total = medida * 1700
                            respuesta = f"El valor estimado para tu barra de largo {medida} cm es: {total:,} $ Pesos.\nSi deseas comunicarte con un asesor, ENVÍA 2."
                            enviar_mensaje(from_number, respuesta)
                            user_steps[from_number] = 'esperando_confirmacion'
                            return jsonify({"status": "barra_ok"})

                        elif step == 'meson_recto_medida':
                            medida = int(text)
                            total = (medida + 100) * 1700
                            respuesta = f"El valor estimado para tu mesón recto es: {total:,} $ Pesos.\nSi deseas comunicarte con un asesor, ENVÍA 2."
                            enviar_mensaje(from_number, respuesta)
                            user_steps[from_number] = 'esperando_confirmacion'
                            return jsonify({"status": "recto_ok"})

                        elif step == 'meson_l_medida':
                            partes = text.replace(" ", "").split("x")
                            if len(partes) == 2:
                                parte1, parte2 = map(int, partes)
                                total = (parte1 + parte2 + 40) * 1700
                                respuesta = f"El valor estimado para tu mesón en L es: {total:,} $ Pesos.\nSi deseas comunicarte con un asesor, ENVÍA 2."
                                enviar_mensaje(from_number, respuesta)
                                user_steps[from_number] = 'esperando_confirmacion'
                                return jsonify({"status": "l_ok"})
                            else:
                                raise ValueError("Formato inválido")

                    except Exception as e:
                        enviar_mensaje(from_number, "Por favor ingresa la medida correctamente. Ej: 150 o 200 x 150")
                        return jsonify({"status": "invalid_measure"})

                    # Consultar reglas desde la base
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute("SELECT respuesta, siguiente_step, tipo, opciones FROM reglas WHERE step = ? AND input_text = ?", (step, text))
                    regla = c.fetchone()
                    conn.close()

                    if regla:
                        respuesta, siguiente, tipo_respuesta, opciones_raw = regla
                        opciones = opciones_raw.split("||") if opciones_raw else []

                        enviar_mensaje(from_number, respuesta, tipo_respuesta=tipo_respuesta, opciones=opciones)
                        if siguiente:
                            user_steps[from_number] = siguiente
                    else:
                        enviar_mensaje(from_number, "Lo siento, no entendí tu respuesta. Por favor intenta nuevamente.")

    return jsonify({"status": "received"})




@app.route('/eliminar_regla/<int:regla_id>', methods=['POST'])
def eliminar_regla(regla_id):
    if "user" not in session or session["rol"] != "admin":
        return redirect(url_for("login"))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM reglas WHERE id = ?", (regla_id,))
    conn.commit()
    conn.close()
    
    return redirect(url_for("configuracion"))


@app.route('/')
def index():
    if "user" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT numero FROM mensajes")
    numeros = [row[0] for row in c.fetchall()]
    chats = []
    for numero in numeros:
        c.execute("SELECT mensaje FROM mensajes WHERE numero = ? ORDER BY timestamp DESC LIMIT 1", (numero,))
        ultimo = c.fetchone()
        requiere_asesor = False
        if ultimo and "asesor" in ultimo[0].lower():
            requiere_asesor = True
        chats.append((numero, requiere_asesor))

    # Leer botones
    c.execute("SELECT id, mensaje FROM botones ORDER BY id")
    botones = c.fetchall()

    conn.close()
    return render_template('index.html', chats=chats, botones=botones)


@app.route('/get_chat/<numero>')
def get_chat(numero):
    if "user" not in session:
        return redirect(url_for("login"))
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT mensaje, tipo, timestamp FROM mensajes WHERE numero = ? ORDER BY timestamp", (numero,))
    mensajes = c.fetchall()
    conn.close()
    return jsonify({'mensajes': mensajes})

@app.route('/send_message', methods=['POST'])
def send_message():
    if "user" not in session:
        return redirect(url_for("login"))
    
    data = request.get_json()
    numero = data.get('numero')
    mensaje = data.get('mensaje')
    enviar_mensaje(numero, mensaje, tipo='asesor')  # <=== importante cambio aquí
    return jsonify({'status': 'success'})

@app.route('/get_chat_list')
def get_chat_list():
    if "user" not in session:
        return redirect(url_for("login"))
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT numero FROM mensajes")
    numeros = [row[0] for row in c.fetchall()]
    chats = []
    for numero in numeros:
        c.execute("SELECT mensaje FROM mensajes WHERE numero = ? ORDER BY timestamp DESC LIMIT 1", (numero,))
        ultimo = c.fetchone()
        requiere_asesor = False
        if ultimo and "asesor" in ultimo[0].lower():
            requiere_asesor = True
        chats.append({'numero': numero, 'asesor': requiere_asesor})
    conn.close()
    return jsonify(chats)


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        hashed = hashlib.sha256(password.encode()).hexdigest()

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT * FROM usuarios WHERE username = ? AND password = ?", (username, hashed))
        user = c.fetchone()
        conn.close()

        if user:
            session["user"] = user[1]
            session["rol"] = user[3]
            return redirect("/")
        else:
            error = "Usuario o contraseña incorrectos"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == '__main__':
    app.run(debug=True)
