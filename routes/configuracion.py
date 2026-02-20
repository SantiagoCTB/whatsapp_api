import importlib.util
import json
import logging
import os
import re
import uuid
from datetime import datetime
from urllib.parse import urlparse, unquote

import requests
from flask import Blueprint, render_template, request, redirect, session, url_for, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

if importlib.util.find_spec("mysql.connector"):
    from mysql.connector import Error as MySQLError
else:  # pragma: no cover - fallback cuando falta el conector
    class MySQLError(Exception):
        pass

from config import Config
from services import tenants
from services.catalog_pdf_worker import enqueue_catalog_pdf_ingest
from services.whatsapp_api import list_phone_numbers
from services.db import get_connection, get_chat_state_definitions

config_bp = Blueprint('configuracion', __name__)
logger = logging.getLogger(__name__)

# El comodín '*' en `input_text` permite avanzar al siguiente paso sin validar
# la respuesta del usuario. Si es la única regla de un paso se ejecuta
# automáticamente; si coexiste con otras, actúa como respuesta por defecto.


def _media_root():
    return tenants.get_media_root()


def _save_followup_media(media_file, media_url):
    if media_file and media_file.filename:
        filename = secure_filename(media_file.filename)
        unique = f"followup_{uuid.uuid4().hex}_{filename}"
        path = os.path.join(_media_root(), unique)
        media_file.save(path)
        return url_for(
            'static',
            filename=tenants.get_uploads_url_path(unique),
            _external=True,
        )
    return media_url


def _require_admin():
    # Debe haber usuario logueado y el rol 'admin' en la lista de roles
    return "user" in session and 'admin' in (session.get('roles') or [])


def _resolve_signup_tenant():
    """Obtiene el tenant activo para el flujo de Embedded Signup.

    - Usa el tenant en contexto cuando existe.
    - Si no hay tenant en contexto, intenta usar el tenant activo (incluye
      DEFAULT_TENANT) y lo recupera del registro.
    """

    tenant = tenants.get_current_tenant()
    if tenant:
        return tenant

    fallback_key = tenants.get_active_tenant_key()
    if not fallback_key:
        return None

    tenant = tenants.get_tenant(fallback_key)
    if tenant:
        return tenant

    # En ambientes donde el DEFAULT_TENANT existe pero aún no se ha
    # materializado en la base, intentamos registrarlo para evitar que el
    # flujo se bloquee por no encontrar el tenant.
    try:
        return tenants.ensure_default_tenant_registered()
    except Exception:
        logger.exception(
            "No se pudo resolver el tenant para Embedded Signup",
            extra={"tenant_key": fallback_key},
        )
        return None


def _fetch_instagram_backfill_counts(tenant):
    if not tenant:
        return None, None

    try:
        tenants.ensure_tenant_schema(tenant)
        conn = get_connection(db_settings=tenant.as_db_settings(), ensure_database=True)
    except MySQLError:
        logger.exception(
            "No se pudo abrir la conexión para el conteo de backfill Instagram",
            extra={"tenant_key": tenant.tenant_key},
        )
        return None, None

    try:
        c = conn.cursor()
        c.execute(
            """
            SELECT COUNT(DISTINCT conversation_id)
              FROM page_messages
             WHERE tenant_key = %s
               AND platform = %s
            """,
            (tenant.tenant_key, "instagram"),
        )
        conversation_count = (c.fetchone() or [0])[0] or 0

        c.execute(
            """
            SELECT COUNT(*)
              FROM page_messages
             WHERE tenant_key = %s
               AND platform = %s
            """,
            (tenant.tenant_key, "instagram"),
        )
        message_count = (c.fetchone() or [0])[0] or 0
    except MySQLError:
        logger.exception(
            "No se pudo obtener el conteo de backfill Instagram",
            extra={"tenant_key": tenant.tenant_key},
        )
        return None, None
    finally:
        conn.close()

    return conversation_count, message_count


def _normalize_input(text):
    """Normaliza una lista separada por comas."""
    return ','.join(t.strip().lower() for t in (text or '').split(',') if t.strip())


def _normalize_platform(value):
    normalized = (value or '').strip().lower()
    if normalized in {'whatsapp', 'messenger', 'instagram'}:
        return normalized
    return None

def _url_ok(url):
    try:
        r = requests.head(url, allow_redirects=True, timeout=5)
        ok = r.status_code == 200
        mime = r.headers.get('Content-Type', '').split(';', 1)[0] if ok else None
        return ok, mime
    except requests.RequestException:
        return False, None




def _graph_get(path: str, access_token: str, *, params: dict | None = None) -> dict:
    if not access_token:
        return {"ok": False, "error": "No se encontró token de acceso para consultar Meta."}

    url = f"https://graph.facebook.com/{Config.FACEBOOK_GRAPH_API_VERSION}/{path.lstrip('/')}"
    query = dict(params or {})
    query["access_token"] = access_token

    try:
        response = requests.get(url, params=query, timeout=20)
    except requests.RequestException:
        logger.warning("No se pudo conectar con Graph API", extra={"path": path})
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    try:
        payload = response.json()
    except ValueError:
        payload = {}

    if response.status_code >= 400:
        return {
            "ok": False,
            "error": "Meta devolvió un error al consultar la información de WhatsApp.",
            "details": payload,
        }

    return {"ok": True, "data": payload}


def _graph_post(path: str, access_token: str, *, data: dict | None = None) -> dict:
    if not access_token:
        return {"ok": False, "error": "No se encontró token de acceso para consultar Meta."}

    url = f"https://graph.facebook.com/{Config.FACEBOOK_GRAPH_API_VERSION}/{path.lstrip('/')}"
    payload = dict(data or {})
    payload["access_token"] = access_token

    try:
        response = requests.post(url, data=payload, timeout=20)
    except requests.RequestException:
        logger.warning("No se pudo conectar con Graph API", extra={"path": path})
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    try:
        body = response.json()
    except ValueError:
        body = {}

    if response.status_code >= 400:
        return {
            "ok": False,
            "error": "Meta devolvió un error al ejecutar la acción sobre WhatsApp.",
            "details": body,
        }

    return {"ok": True, "data": body}


def _resolve_whatsapp_token_and_business_id(tenant) -> tuple[str, str, dict]:
    tenant_env = tenants.get_tenant_env(tenant)
    access_token = (tenant_env.get("META_TOKEN") or tenant_env.get("LONG_LIVED_TOKEN") or "").strip()
    business_id = (tenant_env.get("BUSINESS_ID") or "").strip()
    return access_token, business_id, tenant_env


def _extract_graph_list(payload: dict) -> list:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    if isinstance(data, list):
        return data
    return []


def _whatsapp_phone_number_action(phone_number_id: str, action: str, access_token: str) -> dict:
    normalized_phone_number_id = (phone_number_id or "").strip()
    normalized_action = (action or "").strip().lower()
    if not normalized_phone_number_id:
        return {"ok": False, "error": "Falta PHONE_NUMBER_ID para ejecutar la acción."}
    if normalized_action not in {"register", "deregister"}:
        return {"ok": False, "error": "Acción inválida para el número de WhatsApp."}
    return _graph_post(f"{normalized_phone_number_id}/{normalized_action}", access_token)

def _fetch_page_accounts(user_token: str):
    if not user_token:
        return {"ok": False, "error": "Falta el token de usuario para consultar páginas."}

    url = f"https://graph.facebook.com/{Config.FACEBOOK_GRAPH_API_VERSION}/me/accounts"
    params = {
        "fields": "id,name,access_token",
        "access_token": user_token,
    }

    try:
        response = requests.get(url, params=params, timeout=15)
    except requests.RequestException:
        logger.warning("No se pudo conectar con Graph API para listar páginas.")
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    payload = {}
    try:
        payload = response.json()
    except ValueError:
        payload = {}

    if response.status_code >= 400:
        details = payload.get("error") if isinstance(payload, dict) else None
        return {
            "ok": False,
            "error": "No se pudieron obtener las páginas.",
            "details": details,
        }

    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, list):
        return {"ok": False, "error": "No se encontraron páginas disponibles."}

    pages = []
    for entry in data:
        if not isinstance(entry, dict):
            continue
        page_id = entry.get("id")
        if not page_id:
            continue
        pages.append(
            {
                "id": page_id,
                "name": entry.get("name"),
                "access_token": entry.get("access_token"),
            }
        )

    return {"ok": True, "pages": pages}


def _fetch_page_from_token(page_token: str):
    if not page_token:
        return {"ok": False, "error": "Falta el token de página para continuar."}

    url = f"https://graph.facebook.com/{Config.FACEBOOK_GRAPH_API_VERSION}/me"
    params = {
        "fields": "id,name",
        "access_token": page_token,
    }

    try:
        response = requests.get(url, params=params, timeout=15)
    except requests.RequestException:
        logger.warning("No se pudo conectar con Graph API para consultar la página.")
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    payload = {}
    try:
        payload = response.json()
    except ValueError:
        payload = {}

    if response.status_code >= 400:
        details = payload.get("error") if isinstance(payload, dict) else None
        return {
            "ok": False,
            "error": "No se pudo validar el token de página.",
            "details": details,
        }

    page_id = payload.get("id") if isinstance(payload, dict) else None
    if not page_id:
        return {"ok": False, "error": "No se encontró la página asociada al token."}

    return {
        "ok": True,
        "page": {
            "id": page_id,
            "name": payload.get("name"),
            "access_token": page_token,
        },
    }


def _fetch_instagram_user(user_token: str):
    if not user_token:
        return {"ok": False, "error": "Falta el token de usuario para consultar Instagram."}

    url = f"https://graph.instagram.com/{Config.FACEBOOK_GRAPH_API_VERSION}/me"
    params = {"fields": "user_id,id,username,account_type"}
    headers = {"Authorization": f"Bearer {user_token}"}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=15)
    except requests.RequestException:
        logger.warning("No se pudo conectar con Graph API para consultar Instagram.")
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    payload = {}
    try:
        payload = response.json()
    except ValueError:
        payload = {}

    if response.status_code >= 400:
        details = payload.get("error") if isinstance(payload, dict) else None
        return {
            "ok": False,
            "error": "No se pudo obtener la cuenta de Instagram.",
            "details": details,
        }

    if not isinstance(payload, dict) or not payload.get("id"):
        return {"ok": False, "error": "No se encontró una cuenta de Instagram asociada."}

    return {"ok": True, "account": payload}


def _exchange_instagram_code_for_token(code: str, redirect_uri: str) -> dict:
    if not code:
        return {"ok": False, "error": "Código de autorización vacío."}
    client_id = Config.INSTAGRAM_APP_ID or Config.FACEBOOK_APP_ID
    client_secret = Config.INSTAGRAM_APP_SECRET or Config.FACEBOOK_APP_SECRET
    if not client_id or not client_secret:
        return {
            "ok": False,
            "error": "Falta configurar INSTAGRAM_APP_ID/INSTAGRAM_APP_SECRET o FACEBOOK_APP_ID/FACEBOOK_APP_SECRET.",
        }

    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "authorization_code",
        "redirect_uri": redirect_uri,
        "code": code,
    }

    token_url = (Config.INSTAGRAM_OAUTH_TOKEN_URL or "").strip() or "https://api.instagram.com/oauth/access_token"
    supports_long_lived = "api.instagram.com" in token_url

    try:
        response = requests.post(
            token_url,
            data=payload,
            timeout=15,
        )
    except requests.RequestException:
        logger.warning("No se pudo conectar al endpoint de Instagram OAuth.")
        return {"ok": False, "error": "No se pudo conectar con la API de Instagram."}

    try:
        data = response.json()
    except ValueError:
        data = {}

    if response.status_code >= 400:
        if "api.instagram.com" in token_url:
            fallback_url = (
                f"https://graph.facebook.com/{Config.FACEBOOK_GRAPH_API_VERSION}/oauth/access_token"
            )
            try:
                fallback_response = requests.post(
                    fallback_url,
                    data=payload,
                    timeout=15,
                )
            except requests.RequestException:
                fallback_response = None

            fallback_data = {}
            if fallback_response is not None:
                try:
                    fallback_data = fallback_response.json()
                except ValueError:
                    fallback_data = {}

                if fallback_response.status_code < 400:
                    access_token = fallback_data.get("access_token")
                    if access_token:
                        return {
                            "ok": True,
                            "access_token": access_token,
                            "is_long_lived": False,
                            "raw": fallback_data,
                            "token_url": fallback_url,
                        }

        return {
            "ok": False,
            "error": "No se pudo intercambiar el código de Instagram.",
            "details": data,
            "token_url": token_url,
        }

    access_token = data.get("access_token")
    if not access_token:
        return {"ok": False, "error": "Instagram no devolvió un access_token."}

    if not supports_long_lived:
        return {"ok": True, "access_token": access_token, "is_long_lived": False, "raw": data}

    try:
        long_response = requests.get(
            "https://graph.instagram.com/access_token",
            params={
                "grant_type": "ig_exchange_token",
                "client_secret": client_secret,
                "access_token": access_token,
            },
            timeout=15,
        )
    except requests.RequestException:
        logger.warning("No se pudo conectar al endpoint de token largo de Instagram.")
        return {"ok": True, "access_token": access_token, "is_long_lived": False, "raw": data}

    try:
        long_data = long_response.json()
    except ValueError:
        long_data = {}

    if long_response.status_code >= 400:
        logger.warning(
            "No se pudo obtener el token largo de Instagram.",
            extra={"details": long_data},
        )
        return {"ok": True, "access_token": access_token, "is_long_lived": False, "raw": data}

    long_token = long_data.get("access_token") or access_token
    return {"ok": True, "access_token": long_token, "is_long_lived": True, "raw": long_data}


def _exchange_embedded_signup_code_for_token(code: str, redirect_uri: str | None) -> dict:
    if not code:
        return {"ok": False, "error": "Código de autorización vacío."}
    if not Config.FACEBOOK_APP_ID or not Config.FACEBOOK_APP_SECRET:
        return {
            "ok": False,
            "error": "Falta configurar FACEBOOK_APP_ID o SECRET_PASSWORD_APP.",
        }

    graph_version = (Config.FACEBOOK_GRAPH_API_VERSION or "v24.0").strip() or "v24.0"
    endpoint = f"https://graph.facebook.com/{graph_version}/oauth/access_token"

    params = {
        "client_id": Config.FACEBOOK_APP_ID,
        "client_secret": Config.FACEBOOK_APP_SECRET,
        "code": code,
    }
    if redirect_uri:
        params["redirect_uri"] = redirect_uri

    logger.info(
        "Intercambiando código de Embedded Signup por token (Meta OAuth)",
        extra={
            "graph_endpoint": endpoint,
            "graph_version": graph_version,
            "redirect_uri": redirect_uri,
            "code_prefix": code[:12],
            "code_length": len(code),
        },
    )

    try:
        response = requests.post(
            endpoint,
            data=params,
            timeout=15,
        )
    except requests.RequestException as exc:
        logger.warning(
            "No se pudo conectar al endpoint de token de Meta: %s",
            exc,
            extra={
                "graph_endpoint": endpoint,
                "graph_version": graph_version,
                "redirect_uri": redirect_uri,
            },
        )
        return {"ok": False, "error": "No se pudo conectar con la API de Meta."}

    try:
        data = response.json()
    except ValueError:
        data = {}

    if response.status_code >= 400:
        logger.warning(
            "Meta rechazó el intercambio de código por POST (status=%s): %s",
            response.status_code,
            data,
            extra={
                "graph_endpoint": endpoint,
                "graph_version": graph_version,
                "redirect_uri": redirect_uri,
            },
        )

        try:
            fallback_response = requests.get(
                endpoint,
                params=params,
                timeout=15,
            )
        except requests.RequestException as exc:
            logger.warning("No se pudo ejecutar fallback GET de Meta OAuth: %s", exc)
            fallback_response = None

        if fallback_response is not None:
            try:
                fallback_data = fallback_response.json()
            except ValueError:
                fallback_data = {}

            if fallback_response.status_code < 400:
                access_token = fallback_data.get("access_token")
                if access_token:
                    return {"ok": True, "access_token": access_token, "raw": fallback_data}

            logger.warning(
                "Meta rechazó el intercambio de código por GET fallback (status=%s): %s",
                fallback_response.status_code,
                fallback_data,
                extra={
                    "graph_endpoint": endpoint,
                    "graph_version": graph_version,
                    "redirect_uri": redirect_uri,
                },
            )
            data = fallback_data

        return {
            "ok": False,
            "error": "No se pudo intercambiar el código del Embedded Signup.",
            "details": data,
        }

    access_token = data.get("access_token")
    if not access_token:
        logger.warning(
            "Meta respondió sin access_token en Embedded Signup: %s",
            data,
            extra={
                "graph_endpoint": endpoint,
                "graph_version": graph_version,
                "redirect_uri": redirect_uri,
            },
        )
        return {
            "ok": False,
            "error": "Meta no devolvió un access_token en la respuesta.",
        }

    return {"ok": True, "access_token": access_token, "raw": data}


def _embedded_signup_is_redirect_mismatch(details: dict | None) -> bool:
    if not isinstance(details, dict):
        return False
    error = details.get("error")
    if not isinstance(error, dict):
        return False
    subcode = error.get("error_subcode")
    message = (error.get("message") or "").lower()
    return subcode == 36008 or "redirect_uri" in message


def _build_embedded_signup_error_message(base_error: str | None, details: dict | None) -> str:
    message = (base_error or "No se pudo completar el intercambio del código de Embedded Signup.").strip()
    if not isinstance(details, dict):
        return message

    error = details.get("error")
    if not isinstance(error, dict):
        return message

    code = error.get("code")
    meta_message = (error.get("error_user_msg") or error.get("message") or "").strip()
    if code == 191:
        return (
            "Meta rechazó la URL de redirección porque su dominio no está permitido en la app. "
            "Agrega el dominio exacto en Meta App > Settings > Basic > App Domains y en "
            "Facebook Login/Embedded Signup (Valid OAuth Redirect URIs)."
        )
    if code == 100 and (error.get("error_subcode") == 36008 or "redirect_uri" in meta_message.lower()):
        return (
            "Meta rechazó el código porque el redirect_uri no coincide exactamente con el usado en el diálogo OAuth. "
            "En Embedded Signup con FB.login (SDK), Meta suele usar internamente un redirect_uri de staticxx/facebook; "
            "por eso el servidor intenta primero intercambiar el code sin redirect_uri y luego con variantes. "
            "Verifica que fallback_redirect_uri/URL pública estén registradas en Meta y que no haya slash o dominio distinto."
        )
    if meta_message:
        return f"{message} {meta_message}".strip()
    return message


def _is_probably_local_hostname(hostname: str | None) -> bool:
    if not hostname:
        return True
    lowered = hostname.strip().lower()
    if lowered in {"localhost", "127.0.0.1", "::1", "web"}:
        return True
    return lowered.endswith(".local")


def _build_redirect_uri_attempts(redirect_uri: str | None, fallback_uri: str | None = None) -> list[str]:
    attempts: list[str] = []

    def add_candidate(value: str | None):
        normalized = (value or "").strip()
        if normalized in attempts:
            return
        attempts.append(normalized)

    primary = (redirect_uri or "").strip()
    secondary = (fallback_uri or "").strip()
    configured_whatsapp_redirect = (getattr(Config, "WHATSAPP_OAUTH_REDIRECT_URI", "") or "").strip()

    parsed_primary = urlparse(primary) if primary else None
    parsed_secondary = urlparse(secondary) if secondary else None

    # Primer intento: sin redirect_uri. En FB.login (SDK) Meta puede usar redirect_uri interno (staticxx).
    add_candidate(None)

    # Segundo intento: URI exacta enviada por frontend/backend.
    add_candidate(primary or None)

    # Tercer intento: URI principal de WhatsApp definida en entorno.
    if configured_whatsapp_redirect:
        add_candidate(configured_whatsapp_redirect)

    # Cuarto intento: URI de respaldo explícita, evitando dominios locales accidentales.
    if secondary:
        same_host = bool(parsed_primary and parsed_secondary and parsed_primary.netloc == parsed_secondary.netloc)
        if same_host or not _is_probably_local_hostname(parsed_secondary.hostname):
            add_candidate(secondary)

    # Quinto intento: dominio principal de WHATSAPP_OAUTH_REDIRECT_URI (https://app.whapco.site)
    if configured_whatsapp_redirect:
        parsed_configured = urlparse(configured_whatsapp_redirect)
        if parsed_configured.scheme and parsed_configured.netloc and not _is_probably_local_hostname(parsed_configured.hostname):
            root_domain = f"{parsed_configured.scheme}://{parsed_configured.netloc}"
            add_candidate(root_domain)
            add_candidate(f"{root_domain}/")

    # Variantes con/sin slash final (solo para dominios públicos).
    for candidate in (primary, configured_whatsapp_redirect, secondary):
        normalized = (candidate or "").strip()
        if not normalized:
            continue
        parsed = urlparse(normalized)
        if _is_probably_local_hostname(parsed.hostname):
            continue
        trimmed = normalized.rstrip("/")
        alternate = trimmed if trimmed != normalized else f"{normalized}/"
        add_candidate(alternate)

    return attempts


def _exchange_embedded_signup_code_with_fallbacks(code: str, redirect_uri: str, fallback_uri: str | None = None) -> dict:
    attempts = _build_redirect_uri_attempts(redirect_uri, fallback_uri)

    last_response = None
    for attempt_uri in attempts:
        attempt_label = attempt_uri or "<empty>"
        logger.info(
            "Intentando intercambio de Embedded Signup",
            extra={"redirect_uri": attempt_label, "code_prefix": code[:12], "code_length": len(code)},
        )
        response = _exchange_embedded_signup_code_for_token(code, attempt_uri or None)
        if response.get("ok"):
            return response

        last_response = response
        if not _embedded_signup_is_redirect_mismatch(response.get("details")):
            return response

        logger.warning(
            "Fallo por redirect_uri; se intentará otra variante",
            extra={"redirect_uri": attempt_label, "details": response.get("details")},
        )

    return last_response or {"ok": False, "error": "No se pudo intercambiar el código del Embedded Signup."}


def _handle_instagram_oauth_code(code: str, redirect_uri: str) -> dict:
    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}

    token_response = _exchange_instagram_code_for_token(code, redirect_uri)
    if not token_response.get("ok"):
        return token_response

    access_token = token_response.get("access_token")
    if not access_token:
        return {"ok": False, "error": "No se obtuvo un token de Instagram válido."}

    account_response = _fetch_instagram_user(access_token)
    if not account_response.get("ok"):
        return account_response

    account = account_response.get("account") or {}
    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["INSTAGRAM_TOKEN"] = access_token
    if account.get("id"):
        env_updates["INSTAGRAM_ACCOUNT_ID"] = account.get("id")
    if account.get("user_id") or account.get("id"):
        env_updates["INSTAGRAM_PAGE_ID"] = account.get("user_id") or account.get("id")
    tenants.update_tenant_env(tenant.tenant_key, env_updates)
    tenants.update_tenant_metadata(
        tenant.tenant_key,
        {"instagram_account": account},
    )
    tenants.trigger_page_backfill_for_platform(tenant, "instagram")

    logger.info(
        "Token de Instagram actualizado desde OAuth",
        extra={
            "tenant_key": tenant.tenant_key,
            "instagram_account_id": account.get("id"),
            "instagram_username": account.get("username"),
            "is_long_lived": token_response.get("is_long_lived"),
        },
    )
    return {"ok": True, "account": account, "access_token": access_token}


def _resolve_instagram_redirect_uri(fallback: str) -> str:
    explicit_redirect = (Config.INSTAGRAM_OAUTH_REDIRECT_URI or "").strip()
    if explicit_redirect:
        return explicit_redirect
    signup_url = (Config.SIGNUP_INSTRAGRAM or "").strip()
    if not signup_url:
        return fallback
    try:
        parsed = urlparse(signup_url)
    except ValueError:
        return fallback
    if not parsed.query:
        return fallback
    for entry in parsed.query.split("&"):
        if not entry:
            continue
        key, _, value = entry.partition("=")
        if key == "redirect_uri" and value:
            return value
    return fallback


def _resolve_embedded_signup_redirect_uri(fallback: str) -> str:
    whatsapp_embedded_redirect = (getattr(Config, "WHATSAPP_EMBEDDED_SIGNUP_REDIRECT_URI", "") or "").strip()
    if whatsapp_embedded_redirect:
        return whatsapp_embedded_redirect

    whatsapp_explicit_redirect = (getattr(Config, "WHATSAPP_OAUTH_REDIRECT_URI", "") or "").strip()
    if whatsapp_explicit_redirect:
        return whatsapp_explicit_redirect

    explicit_redirect = (Config.EMBEDDED_SIGNUP_REDIRECT_URI or "").strip()
    if explicit_redirect:
        return explicit_redirect

    signup_url = (Config.SIGNUP_FACEBOOK or "").strip()
    if signup_url:
        try:
            parsed = urlparse(signup_url)
        except ValueError:
            parsed = None
        if parsed and parsed.query:
            for entry in parsed.query.split("&"):
                if not entry:
                    continue
                key, _, value = entry.partition("=")
                if key == "redirect_uri" and value:
                    return unquote(value)

    base_url = (Config.PUBLIC_BASE_URL or "").strip().rstrip("/")
    if base_url:
        return f"{base_url}/configuracion/signup"
    return (fallback or "").rstrip("/")


def _resolve_page_user_token(platform: str | None, tenant_env: dict, provided_token: str) -> str:
    normalized = (platform or "").strip().lower()
    if normalized == "instagram":
        return (
            provided_token
            or (tenant_env.get("INSTAGRAM_TOKEN") or "").strip()
            or (tenant_env.get("MESSENGER_TOKEN") or "").strip()
        )
    return provided_token or (tenant_env.get("MESSENGER_TOKEN") or "").strip()


def _resolve_user_token_key(platform: str | None) -> str:
    normalized = (platform or "").strip().lower()
    if normalized == "instagram":
        return "INSTAGRAM_TOKEN"
    return "MESSENGER_TOKEN"


def _resolve_page_env_key(platform: str, key: str) -> str:
    normalized = (platform or "").strip().lower()
    if normalized == "instagram":
        return f"INSTAGRAM_{key}"
    return f"MESSENGER_{key}"


def _resolve_page_env_value(platform: str, tenant_env: dict) -> str:
    normalized = (platform or "").strip().lower()
    page_id = (tenant_env.get(_resolve_page_env_key(normalized, "PAGE_ID")) or "").strip()
    if page_id:
        return page_id
    legacy_platform = (tenant_env.get("PLATFORM") or "").strip().lower()
    if legacy_platform == normalized:
        return (tenant_env.get("PAGE_ID") or "").strip()
    return ""


def _normalize_page_selection(metadata: dict | None) -> dict:
    selection = {}
    if not isinstance(metadata, dict):
        return selection

    raw = metadata.get("page_selection")
    if not isinstance(raw, dict):
        return selection

    if "messenger" in raw or "instagram" in raw:
        for platform in ("messenger", "instagram"):
            entry = raw.get(platform)
            if isinstance(entry, dict) and entry.get("page_id"):
                selection[platform] = {
                    "page_id": entry.get("page_id"),
                    "page_name": entry.get("page_name"),
                }
        return selection

    platform = (raw.get("platform") or "").strip().lower()
    if platform in {"messenger", "instagram"} and raw.get("page_id"):
        selection[platform] = {
            "page_id": raw.get("page_id"),
            "page_name": raw.get("page_name"),
        }

    return selection


def _normalize_state_key(raw_key: str | None) -> str | None:
    if not raw_key:
        return None
    cleaned = re.sub(r"[^a-z0-9_-]+", "_", raw_key.strip().lower()).strip("_")
    if not cleaned:
        return None
    return cleaned[:40]


def _coerce_hex_color(value: str | None, default: str) -> str:
    if not value:
        return default
    candidate = value.strip()
    if not candidate:
        return default
    if not candidate.startswith("#"):
        candidate = f"#{candidate}"
    if re.fullmatch(r"#[0-9a-fA-F]{6}", candidate):
        return candidate
    return default


def _ensure_ia_config_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS ia_config (
            id INT AUTO_INCREMENT PRIMARY KEY,
            model_name VARCHAR(100) NOT NULL DEFAULT 'o4-mini',
            model_token TEXT NULL,
            system_prompt TEXT NULL,
            business_description TEXT NULL,
            followup_message_1 TEXT NULL,
            followup_message_2 TEXT NULL,
            followup_message_3 TEXT NULL,
            followup_media_url_1 TEXT NULL,
            followup_media_url_2 TEXT NULL,
            followup_media_url_3 TEXT NULL,
            followup_media_tipo_1 VARCHAR(20) NULL,
            followup_media_tipo_2 VARCHAR(20) NULL,
            followup_media_tipo_3 VARCHAR(20) NULL,
            followup_interval_minutes INT NULL,
            enabled TINYINT(1) NOT NULL DEFAULT 1,
            pdf_filename VARCHAR(255) NULL,
            pdf_original_name VARCHAR(255) NULL,
            pdf_mime VARCHAR(100) NULL,
            pdf_size BIGINT NULL,
            pdf_uploaded_at DATETIME NULL,
            pdf_source_url TEXT NULL,
            pdf_ingest_state VARCHAR(20) NULL,
            pdf_ingest_started_at DATETIME NULL,
            pdf_ingest_finished_at DATETIME NULL,
            pdf_ingest_error TEXT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB;
        """
    )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'enabled';")
    has_enabled = cursor.fetchone() is not None
    if not has_enabled:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN enabled TINYINT(1) NOT NULL DEFAULT 1 AFTER model_token;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'system_prompt';")
    has_system_prompt = cursor.fetchone() is not None
    if not has_system_prompt:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN system_prompt TEXT NULL AFTER model_token;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'business_description';")
    has_business_description = cursor.fetchone() is not None
    if not has_business_description:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN business_description TEXT NULL AFTER system_prompt;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_message_1';")
    has_followup_message_1 = cursor.fetchone() is not None
    if not has_followup_message_1:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_message_1 TEXT NULL AFTER business_description;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_message_2';")
    has_followup_message_2 = cursor.fetchone() is not None
    if not has_followup_message_2:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_message_2 TEXT NULL AFTER followup_message_1;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_message_3';")
    has_followup_message_3 = cursor.fetchone() is not None
    if not has_followup_message_3:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_message_3 TEXT NULL AFTER followup_message_2;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_interval_minutes';")
    has_followup_interval = cursor.fetchone() is not None
    if not has_followup_interval:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_interval_minutes INT NULL AFTER followup_message_3;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_url_1';")
    has_followup_media_url_1 = cursor.fetchone() is not None
    if not has_followup_media_url_1:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_url_1 TEXT NULL AFTER followup_interval_minutes;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_url_2';")
    has_followup_media_url_2 = cursor.fetchone() is not None
    if not has_followup_media_url_2:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_url_2 TEXT NULL AFTER followup_media_url_1;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_url_3';")
    has_followup_media_url_3 = cursor.fetchone() is not None
    if not has_followup_media_url_3:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_url_3 TEXT NULL AFTER followup_media_url_2;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_tipo_1';")
    has_followup_media_tipo_1 = cursor.fetchone() is not None
    if not has_followup_media_tipo_1:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_tipo_1 VARCHAR(20) NULL AFTER followup_media_url_3;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_tipo_2';")
    has_followup_media_tipo_2 = cursor.fetchone() is not None
    if not has_followup_media_tipo_2:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_tipo_2 VARCHAR(20) NULL AFTER followup_media_tipo_1;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'followup_media_tipo_3';")
    has_followup_media_tipo_3 = cursor.fetchone() is not None
    if not has_followup_media_tipo_3:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN followup_media_tipo_3 VARCHAR(20) NULL AFTER followup_media_tipo_2;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'pdf_source_url';")
    has_source_url = cursor.fetchone() is not None
    if not has_source_url:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN pdf_source_url TEXT NULL AFTER pdf_uploaded_at;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'pdf_ingest_state';")
    has_ingest_state = cursor.fetchone() is not None
    if not has_ingest_state:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN pdf_ingest_state VARCHAR(20) NULL AFTER pdf_source_url;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'pdf_ingest_started_at';")
    has_ingest_started = cursor.fetchone() is not None
    if not has_ingest_started:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN pdf_ingest_started_at DATETIME NULL AFTER pdf_ingest_state;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'pdf_ingest_finished_at';")
    has_ingest_finished = cursor.fetchone() is not None
    if not has_ingest_finished:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN pdf_ingest_finished_at DATETIME NULL AFTER pdf_ingest_started_at;"
        )

    cursor.execute("SHOW COLUMNS FROM ia_config LIKE 'pdf_ingest_error';")
    has_ingest_error = cursor.fetchone() is not None
    if not has_ingest_error:
        cursor.execute(
            "ALTER TABLE ia_config ADD COLUMN pdf_ingest_error TEXT NULL AFTER pdf_ingest_finished_at;"
        )


def _get_ia_config(cursor):
    keys = [
        "id",
        "model_name",
        "model_token",
        "enabled",
        "system_prompt",
        "business_description",
        "followup_message_1",
        "followup_message_2",
        "followup_message_3",
        "followup_interval_minutes",
        "followup_media_url_1",
        "followup_media_url_2",
        "followup_media_url_3",
        "followup_media_tipo_1",
        "followup_media_tipo_2",
        "followup_media_tipo_3",
        "pdf_filename",
        "pdf_original_name",
        "pdf_mime",
        "pdf_size",
        "pdf_uploaded_at",
        "pdf_source_url",
        "pdf_ingest_state",
        "pdf_ingest_started_at",
        "pdf_ingest_finished_at",
        "pdf_ingest_error",
    ]

    queries = [
        (
            """
            SELECT id, model_name, model_token, enabled, system_prompt, business_description,
                   followup_message_1, followup_message_2, followup_message_3, followup_interval_minutes,
                   followup_media_url_1, followup_media_url_2, followup_media_url_3,
                   followup_media_tipo_1, followup_media_tipo_2, followup_media_tipo_3,
                   pdf_filename, pdf_original_name, pdf_mime, pdf_size, pdf_uploaded_at,
                   pdf_source_url, pdf_ingest_state, pdf_ingest_started_at, pdf_ingest_finished_at,
                   pdf_ingest_error
              FROM ia_config
          ORDER BY id DESC
             LIMIT 1
            """,
            keys,
        ),
        (
            """
            SELECT id, model_name, model_token, enabled, system_prompt, business_description,
                   pdf_filename, pdf_original_name, pdf_mime, pdf_size, pdf_uploaded_at,
                   pdf_source_url, pdf_ingest_state, pdf_ingest_started_at, pdf_ingest_finished_at,
                   pdf_ingest_error
              FROM ia_config
          ORDER BY id DESC
             LIMIT 1
            """,
            [
                "id",
                "model_name",
                "model_token",
                "enabled",
                "system_prompt",
                "business_description",
                "pdf_filename",
                "pdf_original_name",
                "pdf_mime",
                "pdf_size",
                "pdf_uploaded_at",
                "pdf_source_url",
                "pdf_ingest_state",
                "pdf_ingest_started_at",
                "pdf_ingest_finished_at",
                "pdf_ingest_error",
            ],
        ),
        (
            """
            SELECT id, model_name, model_token, pdf_filename, pdf_original_name,
                   pdf_mime, pdf_size, pdf_uploaded_at, pdf_source_url,
                   pdf_ingest_state, pdf_ingest_started_at, pdf_ingest_finished_at,
                   pdf_ingest_error
              FROM ia_config
          ORDER BY id DESC
             LIMIT 1
            """,
            [
                "id",
                "model_name",
                "model_token",
                "pdf_filename",
                "pdf_original_name",
                "pdf_mime",
                "pdf_size",
                "pdf_uploaded_at",
                "pdf_source_url",
                "pdf_ingest_state",
                "pdf_ingest_started_at",
                "pdf_ingest_finished_at",
                "pdf_ingest_error",
            ],
        ),
    ]

    row = None
    row_keys = None
    for query, query_keys in queries:
        try:
            cursor.execute(query)
            row = cursor.fetchone()
            row_keys = query_keys
            break
        except Exception:
            continue

    if not row:
        return None

    data = {key: None for key in keys}
    for key, value in zip(row_keys or [], row):
        data[key] = value

    if data["enabled"] is None:
        data["enabled"] = 1

    return data


def _botones_opciones_column(c, conn):
    """Asegura que la columna ``opciones`` exista en ``botones``.

    Devuelve la expresión SQL a utilizar en el SELECT para soportar
    instalaciones antiguas donde aún no existe la columna. En esos casos
    se intentará crearla y, si no es posible, se regresa ``NULL`` como
    marcador para evitar errores ``Unknown column``.
    """

    has_opciones = True
    try:
        c.execute("SHOW COLUMNS FROM botones LIKE 'opciones';")
        has_opciones = c.fetchone() is not None
        if not has_opciones:
            try:
                c.execute("ALTER TABLE botones ADD COLUMN opciones TEXT NULL;")
                conn.commit()
                has_opciones = True
            except MySQLError:
                conn.rollback()
                has_opciones = False
    except MySQLError:
        has_opciones = False

    return "b.opciones" if has_opciones else "NULL AS opciones"


def _botones_categoria_column(c, conn):
    """Asegura que la columna ``categoria`` exista en ``botones``.

    Devuelve la expresión SQL a utilizar en el SELECT para soportar
    instalaciones antiguas donde aún no existe la columna.
    """

    has_categoria = True
    try:
        c.execute("SHOW COLUMNS FROM botones LIKE 'categoria';")
        has_categoria = c.fetchone() is not None
        if not has_categoria:
            try:
                c.execute("ALTER TABLE botones ADD COLUMN categoria VARCHAR(100) NULL;")
                conn.commit()
                has_categoria = True
            except MySQLError:
                conn.rollback()
                has_categoria = False
    except MySQLError:
        has_categoria = False

    return "b.categoria" if has_categoria else "NULL AS categoria"


def _selected_user_ids(form, cursor, default_to_session=True):
    if form.get("clear_users") == "1":
        return []
    user_ids = []
    for value in form.getlist("user_ids"):
        if str(value).isdigit():
            user_ids.append(int(value))

    if not user_ids and default_to_session:
        username = session.get("user")
        if username:
            cursor.execute("SELECT id FROM usuarios WHERE username = %s", (username,))
            row = cursor.fetchone()
            if row:
                user_ids = [row[0]]

    return user_ids


def _assign_boton_users(cursor, boton_id, user_ids):
    if not user_ids:
        return
    for user_id in user_ids:
        cursor.execute(
            "INSERT IGNORE INTO boton_usuarios (boton_id, user_id) VALUES (%s, %s)",
            (boton_id, user_id),
        )

def _reglas_view(template_name):
    """Renderiza las vistas de reglas.
    El comodín '*' en `input_text` avanza al siguiente paso sin validar
    la respuesta del usuario; si existen otras reglas, actúa como opción
    por defecto."""
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        # --- Migraciones defensivas de nuevas columnas ---
        c.execute("SHOW COLUMNS FROM reglas LIKE 'rol_keyword';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN rol_keyword VARCHAR(20) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'calculo';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN calculo TEXT NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'handler';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN handler VARCHAR(50) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'platform';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN platform VARCHAR(20) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'media_url';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN media_url TEXT NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'media_tipo';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN media_tipo VARCHAR(20) NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'active_hours';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN active_hours TEXT NULL;")
            conn.commit()
        c.execute("SHOW COLUMNS FROM reglas LIKE 'active_days';")
        if not c.fetchone():
            c.execute("ALTER TABLE reglas ADD COLUMN active_days TEXT NULL;")
            conn.commit()

        if request.method == 'POST':
            # Importar desde Excel
            if 'archivo' in request.files and request.files['archivo']:
                archivo = request.files['archivo']
                wb = load_workbook(archivo)
                hoja = wb.active
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    if not fila:
                        continue
                    # Permitir archivos con columnas opcionales
                    datos = list(fila) + [None] * 14
                    (
                        step,
                        input_text,
                        respuesta,
                        siguiente_step,
                        tipo,
                        media_url,
                        media_tipo,
                        opciones,
                        rol_keyword,
                        calculo,
                        handler,
                        platform_raw,
                        active_hours,
                        active_days,
                    ) = datos[:14]
                    url_ok = False
                    detected_type = None
                    if media_url:
                        url_ok, detected_type = _url_ok(str(media_url))
                        if not url_ok:
                            media_url = None
                            media_tipo = None
                        else:
                            media_tipo = media_tipo or detected_type
                    if media_tipo:
                        media_tipo = str(media_tipo).split(';', 1)[0]
                    # Normalizar campos clave
                    step = (step or '').strip().lower()
                    input_text = _normalize_input(input_text)
                    siguiente_step = _normalize_input(siguiente_step) or None
                    platform = _normalize_platform(platform_raw)
                    active_hours = (active_hours or '').strip() or None
                    active_days = (active_days or '').strip() or None

                    c.execute(
                        """
                        SELECT id FROM reglas
                         WHERE step = %s AND input_text = %s
                           AND (
                               platform = %s
                               OR (%s IS NULL AND (platform IS NULL OR platform = ''))
                           )
                        """,
                        (step, input_text, platform, platform),
                    )
                    existente = c.fetchone()
                    if existente:
                        regla_id = existente[0]
                        c.execute(
                            """
                            UPDATE reglas
                               SET respuesta = %s,
                                   siguiente_step = %s,
                                   platform = %s,
                                   tipo = %s,
                                   media_url = %s,
                                   media_tipo = %s,
                                   opciones = %s,
                                   rol_keyword = %s,
                                   calculo = %s,
                                   handler = %s,
                                   active_hours = %s,
                                   active_days = %s
                             WHERE id = %s
                            """,
                            (
                                respuesta,
                                siguiente_step,
                                platform,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                                active_hours,
                                active_days,
                                regla_id,
                            ),
                        )
                        c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                        if media_url and url_ok:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, media_url, media_tipo),
                            )
                    else:
                        c.execute(
                            "INSERT INTO reglas (step, input_text, respuesta, siguiente_step, platform, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler, active_hours, active_days) "
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                            (
                                step,
                                input_text,
                                respuesta,
                                siguiente_step,
                                platform,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                                active_hours,
                                active_days,
                            ),
                        )
                        regla_id = c.lastrowid
                        if media_url and url_ok:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, media_url, media_tipo),
                            )
                conn.commit()
            else:
                # Entrada manual desde formulario
                step = (request.form['step'] or '').strip().lower() or None
                input_text = _normalize_input(request.form['input_text']) or None
                respuesta = request.form['respuesta']
                siguiente_step = _normalize_input(request.form.get('siguiente_step')) or None
                tipo = request.form.get('tipo', 'texto')
                media_files = request.files.getlist('media') or request.files.getlist('media[]')
                media_url_field = request.form.get('media_url')
                medias = []
                for media_file in media_files:
                    if media_file and media_file.filename:
                        filename = secure_filename(media_file.filename)
                        unique = f"{uuid.uuid4().hex}_{filename}"
                        path = os.path.join(_media_root(), unique)
                        media_file.save(path)
                        url = url_for(
                            'static',
                            filename=tenants.get_uploads_url_path(unique),
                            _external=True,
                        )
                        medias.append((url, media_file.mimetype.split(';', 1)[0]))
                if media_url_field:
                    for url in [u.strip() for u in re.split(r'[\n,]+', media_url_field) if u.strip()]:
                        ok, content_type = _url_ok(url)
                        if not ok:
                            return f"URL no válida: {url}", 400
                        medias.append((url, content_type))
                media_url = medias[0][0] if medias else None
                media_tipo = medias[0][1] if medias else None
                opciones = request.form['opciones']
                list_header = request.form.get('list_header')
                list_footer = request.form.get('list_footer')
                list_button = request.form.get('list_button')
                sections_raw = request.form.get('sections')
                if tipo == 'lista':
                    if not opciones:
                        try:
                            sections = json.loads(sections_raw) if sections_raw else []
                        except Exception:
                            sections = []
                        opts = {
                            'header': list_header,
                            'footer': list_footer,
                            'button': list_button,
                            'sections': sections
                        }
                        opciones = json.dumps(opts)
                elif tipo == 'flow':
                    opciones_raw = (request.form.get('opciones') or '').strip()
                    flow_payload = {}
                    flow_keys = [k for k in request.form.keys() if k.startswith('flow_')]
                    for key in flow_keys:
                        value = request.form.get(key)
                        if key in {'flow_payload', 'flow_data'} and value:
                            try:
                                flow_payload[key] = json.loads(value)
                            except Exception:
                                flow_payload[key] = value
                        else:
                            flow_payload[key] = value
                    if flow_payload:
                        try:
                            opciones = json.dumps(flow_payload, ensure_ascii=False)
                        except (TypeError, ValueError):
                            opciones = json.dumps({k: str(v) if v is not None else '' for k, v in flow_payload.items()}, ensure_ascii=False)
                    elif opciones_raw:
                        try:
                            opciones = json.dumps(json.loads(opciones_raw), ensure_ascii=False)
                        except Exception:
                            opciones = opciones_raw
                    else:
                        opciones = ''
                rol_keyword = request.form.get('rol_keyword')
                calculo = request.form.get('calculo')
                handler = request.form.get('handler')
                active_hours = (request.form.get('active_hours') or '').strip() or None
                active_days = (request.form.get('active_days') or '').strip() or None
                regla_id = request.form.get('regla_id')
                platform = _normalize_platform(request.form.get('platform'))

                if regla_id:
                    c.execute(
                        """
                        UPDATE reglas
                           SET step = %s,
                               input_text = %s,
                               respuesta = %s,
                               siguiente_step = %s,
                               platform = %s,
                               tipo = %s,
                               media_url = %s,
                               media_tipo = %s,
                               opciones = %s,
                               rol_keyword = %s,
                               calculo = %s,
                               handler = %s,
                               active_hours = %s,
                               active_days = %s
                         WHERE id = %s
                        """,
                        (
                            step,
                            input_text,
                            respuesta,
                            siguiente_step,
                            platform,
                            tipo,
                            media_url,
                            media_tipo,
                            opciones,
                            rol_keyword,
                            calculo,
                            handler,
                            active_hours,
                            active_days,
                            regla_id,
                        ),
                    )
                    c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                    for url, tipo_media in medias:
                        c.execute(
                            "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                            (regla_id, url, tipo_media),
                        )
                else:
                    c.execute(
                        """
                        SELECT id FROM reglas
                         WHERE step = %s AND input_text = %s
                           AND (
                               platform = %s
                               OR (%s IS NULL AND (platform IS NULL OR platform = ''))
                           )
                        """,
                        (step, input_text, platform, platform),
                    )
                    existente = c.fetchone()
                    if existente:
                        regla_id = existente[0]
                        c.execute(
                            """
                            UPDATE reglas
                               SET respuesta = %s,
                                   siguiente_step = %s,
                                   platform = %s,
                                   tipo = %s,
                                   media_url = %s,
                                   media_tipo = %s,
                                   opciones = %s,
                                   rol_keyword = %s,
                                   calculo = %s,
                                   handler = %s,
                                   active_hours = %s,
                                   active_days = %s
                             WHERE id = %s
                            """,
                            (
                                respuesta,
                                siguiente_step,
                                platform,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                                active_hours,
                                active_days,
                                regla_id,
                            ),
                        )
                        c.execute("DELETE FROM regla_medias WHERE regla_id=%s", (regla_id,))
                        for url, tipo_media in medias:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, url, tipo_media),
                            )
                    else:
                        c.execute(
                            "INSERT INTO reglas (step, input_text, respuesta, siguiente_step, platform, tipo, media_url, media_tipo, opciones, rol_keyword, calculo, handler, active_hours, active_days) "
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                            (
                                step,
                                input_text,
                                respuesta,
                                siguiente_step,
                                platform,
                                tipo,
                                media_url,
                                media_tipo,
                                opciones,
                                rol_keyword,
                                calculo,
                                handler,
                                active_hours,
                                active_days,
                            ),
                        )
                        regla_id = c.lastrowid
                        for url, tipo_media in medias:
                            c.execute(
                                "INSERT INTO regla_medias (regla_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (regla_id, url, tipo_media),
                            )
                conn.commit()

        # Listar todas las reglas
        c.execute(
            """
            SELECT r.id, r.step, r.input_text, r.respuesta, r.siguiente_step, r.platform, r.tipo,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos,
                   r.opciones, r.rol_keyword, r.calculo, r.handler,
                   r.active_hours, r.active_days
              FROM reglas r
              LEFT JOIN regla_medias m ON r.id = m.regla_id
             GROUP BY r.id
             ORDER BY r.id DESC
            """
        )
        rows = c.fetchall()
        reglas = []
        for row in rows:
            d = {
                'id': row[0],
                'step': row[1],
                'input_text': row[2],
                'respuesta': row[3],
                'siguiente_step': row[4],
                'platform': row[5],
                'tipo': row[6],
                'media_urls': (row[7] or '').split('||') if row[7] else [],
                'media_tipos': (row[8] or '').split('||') if row[8] else [],
                'opciones': row[9] or '',
                'rol_keyword': row[10],
                'calculo': row[11],
                'handler': row[12],
                'active_hours': row[13],
                'active_days': row[14],
                'header': None,
                'button': None,
                'footer': None,
                'flow': None,
                'opciones_pretty': None,
            }
            if d['opciones']:
                parsed_opts = None
                try:
                    parsed_opts = json.loads(d['opciones'])
                except Exception:
                    parsed_opts = None

                if d['tipo'] == 'lista' and isinstance(parsed_opts, dict):
                    d['header'] = parsed_opts.get('header')
                    d['button'] = parsed_opts.get('button')
                    d['footer'] = parsed_opts.get('footer')
                elif d['tipo'] == 'flow' and isinstance(parsed_opts, dict):
                    for key, value in parsed_opts.items():
                        if isinstance(value, (dict, list)):
                            try:
                                d[key] = json.dumps(value, ensure_ascii=False)
                            except (TypeError, ValueError):
                                d[key] = value
                        else:
                            d[key] = value
                    d['flow_options'] = parsed_opts
            reglas.append(d)
        c.execute("SELECT name, keyword FROM roles ORDER BY name")
        roles_rows = c.fetchall()
        roles = [{"name": row[0], "keyword": row[1]} for row in roles_rows]
        tenant_env = dict(tenants.get_current_tenant_env() or {})
        instagram_token_present = bool((tenant_env.get("INSTAGRAM_TOKEN") or "").strip())
        chat_state_definitions = get_chat_state_definitions(include_hidden=True)
        return render_template(
            template_name,
            reglas=reglas,
            roles=roles,
            chat_state_definitions=chat_state_definitions,
            instagram_token_present=instagram_token_present,
        )
    finally:
        conn.close()


@config_bp.route('/chat_states', methods=['POST'])
def save_chat_state_definition():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    get_chat_state_definitions(include_hidden=True)
    raw_key = request.form.get('state_key')
    original_key = request.form.get('original_key')
    label = (request.form.get('label') or '').strip()
    color_hex = _coerce_hex_color(request.form.get('color_hex'), '#666666')
    text_color_hex = _coerce_hex_color(request.form.get('text_color_hex'), '#ffffff')
    priority_raw = request.form.get('priority')
    visible = 1 if request.form.get('visible') in {'1', 'true', 'on', 'yes'} else 0

    state_key = _normalize_state_key(raw_key)
    if not state_key:
        return redirect(url_for('configuracion.reglas'))

    if not label:
        label = state_key.replace("_", " ").title()

    try:
        priority = int(priority_raw) if priority_raw is not None else 0
    except (TypeError, ValueError):
        priority = 0

    conn = get_connection()
    c = conn.cursor()
    try:
        if original_key:
            original_key = _normalize_state_key(original_key)
        if original_key and original_key != state_key:
            c.execute(
                "DELETE FROM chat_state_definitions WHERE state_key = %s",
                (original_key,),
            )

        c.execute(
            """
            INSERT INTO chat_state_definitions
                (state_key, label, color_hex, text_color_hex, priority, visible)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                label = VALUES(label),
                color_hex = VALUES(color_hex),
                text_color_hex = VALUES(text_color_hex),
                priority = VALUES(priority),
                visible = VALUES(visible)
            """,
            (state_key, label, color_hex, text_color_hex, priority, visible),
        )
        conn.commit()
    finally:
        conn.close()

    return redirect(url_for('configuracion.reglas'))


@config_bp.route('/chat_states/delete', methods=['POST'])
def delete_chat_state_definition():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    state_key = _normalize_state_key(request.form.get('state_key'))
    if not state_key:
        return redirect(url_for('configuracion.reglas'))

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute(
            "DELETE FROM chat_state_definitions WHERE state_key = %s",
            (state_key,),
        )
        conn.commit()
    finally:
        conn.close()

    return redirect(url_for('configuracion.reglas'))


@config_bp.route('/configuracion/ia', methods=['GET', 'POST'])
def configuracion_ia():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    status_message = None
    error_message = None
    try:
        _ensure_ia_config_table(c)
        conn.commit()

        ia_config = _get_ia_config(c)
        pdf_url = None
        if ia_config and ia_config.get('pdf_filename'):
            pdf_filename = ia_config['pdf_filename']
            preferred_path = os.path.join(_media_root(), pdf_filename)
            preferred_url_path = tenants.get_uploads_url_path(pdf_filename)
            if not os.path.exists(preferred_path):
                legacy_path = os.path.join(_media_root(), 'ia', pdf_filename)
                if os.path.exists(legacy_path):
                    preferred_url_path = tenants.get_uploads_url_path(f"ia/{pdf_filename}")
            pdf_url = url_for('static', filename=preferred_url_path)

        if request.method == 'POST':
            ia_model = (request.form.get('ia_model') or 'o4-mini').strip() or 'o4-mini'
            ia_token = (request.form.get('ia_token') or '').strip()
            ia_enabled = 1 if request.form.get('ia_enabled') in {'on', '1', 'true', 't'} else 0
            system_prompt = (request.form.get('system_prompt') or '').strip() or None
            business_description = (
                (request.form.get('business_description') or '').strip() or None
            )
            followup_message_1 = (request.form.get('followup_message_1') or '').strip() or None
            followup_message_2 = (request.form.get('followup_message_2') or '').strip() or None
            followup_message_3 = (request.form.get('followup_message_3') or '').strip() or None
            followup_media_url_1 = (request.form.get('followup_media_url_1') or '').strip() or None
            followup_media_url_2 = (request.form.get('followup_media_url_2') or '').strip() or None
            followup_media_url_3 = (request.form.get('followup_media_url_3') or '').strip() or None
            followup_media_file_1 = request.files.get('followup_media_file_1')
            followup_media_file_2 = request.files.get('followup_media_file_2')
            followup_media_file_3 = request.files.get('followup_media_file_3')
            followup_media_tipo_1 = (request.form.get('followup_media_tipo_1') or '').strip() or None
            followup_media_tipo_2 = (request.form.get('followup_media_tipo_2') or '').strip() or None
            followup_media_tipo_3 = (request.form.get('followup_media_tipo_3') or '').strip() or None
            followup_interval_minutes = request.form.get('followup_interval_minutes')
            try:
                followup_interval_minutes = int(followup_interval_minutes)
            except (TypeError, ValueError):
                followup_interval_minutes = None
            if followup_interval_minutes is not None and followup_interval_minutes < 0:
                followup_interval_minutes = 0
            catalog_url = (request.form.get('catalogo_url') or '').strip()
            pdf_file = request.files.get('catalogo_pdf')
            pdf_dir = _media_root()
            os.makedirs(pdf_dir, exist_ok=True)
            stored_catalog_name = 'catalogo.pdf'

            new_pdf = None
            old_pdf_path = None

            if not ia_token:
                error_message = 'El token del modelo es obligatorio.'

            if pdf_file and pdf_file.filename and catalog_url:
                error_message = 'Sube un PDF o indica una URL, pero no ambas opciones.'

            if not error_message:
                followup_pairs = [
                    (1, followup_media_file_1, followup_media_url_1),
                    (2, followup_media_file_2, followup_media_url_2),
                    (3, followup_media_file_3, followup_media_url_3),
                ]
                for idx, media_file, media_url in followup_pairs:
                    if media_file and media_file.filename and media_url:
                        error_message = (
                            f'Sube un archivo o indica una URL en el follow-up {idx},'
                            ' pero no ambas opciones.'
                        )
                        break

            if pdf_file and pdf_file.filename and not error_message:
                filename = secure_filename(pdf_file.filename)
                mime = (pdf_file.mimetype or '').lower()
                if not filename.lower().endswith('.pdf'):
                    error_message = 'Solo se permiten archivos PDF.'
                elif mime and 'pdf' not in mime:
                    error_message = 'El archivo subido no parece ser un PDF válido.'
                else:
                    stored_name = stored_catalog_name
                    path = os.path.join(pdf_dir, stored_name)
                    pdf_file.save(path)
                    pdf_size = os.path.getsize(path)
                    new_pdf = {
                        'stored_name': stored_name,
                        'original_name': filename,
                        'mime': pdf_file.mimetype or 'application/pdf',
                        'size': pdf_size,
                        'source_url': None,
                    }
                    if ia_config and ia_config.get('pdf_filename'):
                        old_pdf_path = os.path.join(pdf_dir, ia_config['pdf_filename'])
                        if not os.path.exists(old_pdf_path):
                            legacy_path = os.path.join(pdf_dir, 'ia', ia_config['pdf_filename'])
                            if os.path.exists(legacy_path):
                                old_pdf_path = legacy_path

            elif catalog_url and not error_message:
                ok, mime = _url_ok(catalog_url)
                if not ok:
                    error_message = 'La URL del catálogo no está disponible o respondió con error.'
                elif mime and 'pdf' not in mime.lower():
                    error_message = 'La URL no apunta a un PDF válido.'
                else:
                    parsed = urlparse(catalog_url)
                    base_name = os.path.basename(parsed.path) or 'catalogo.pdf'
                    filename = secure_filename(base_name) or 'catalogo.pdf'
                    stored_name = stored_catalog_name
                    path = os.path.join(pdf_dir, stored_name)
                    try:
                        with requests.get(catalog_url, stream=True, timeout=120) as resp:
                            if resp.status_code != 200:
                                error_message = 'No se pudo descargar el catálogo desde la URL proporcionada.'
                            else:
                                with open(path, 'wb') as fh:
                                    for chunk in resp.iter_content(chunk_size=8192):
                                        if not chunk:
                                            continue
                                        fh.write(chunk)

                                if not error_message:
                                    pdf_size = os.path.getsize(path)
                                    new_pdf = {
                                        'stored_name': stored_name,
                                        'original_name': filename,
                                        'mime': resp.headers.get('Content-Type', 'application/pdf'),
                                        'size': pdf_size,
                                        'source_url': catalog_url,
                                    }
                                    if ia_config and ia_config.get('pdf_filename'):
                                        old_pdf_path = os.path.join(pdf_dir, ia_config['pdf_filename'])
                                        if not os.path.exists(old_pdf_path):
                                            legacy_path = os.path.join(pdf_dir, 'ia', ia_config['pdf_filename'])
                                            if os.path.exists(legacy_path):
                                                old_pdf_path = legacy_path

                    except requests.RequestException:
                        error_message = 'No se pudo descargar el catálogo desde la URL proporcionada.'
                    if error_message and os.path.exists(path):
                        try:
                            os.remove(path)
                        except OSError:
                            pass

            if not error_message:
                os.makedirs(_media_root(), exist_ok=True)
                followup_media_url_1 = _save_followup_media(
                    followup_media_file_1, followup_media_url_1
                )
                followup_media_url_2 = _save_followup_media(
                    followup_media_file_2, followup_media_url_2
                )
                followup_media_url_3 = _save_followup_media(
                    followup_media_file_3, followup_media_url_3
                )

            if not error_message:
                ingest_state = ia_config.get("pdf_ingest_state") if ia_config else None
                ingest_started_at = (
                    ia_config.get("pdf_ingest_started_at") if ia_config else None
                )
                ingest_finished_at = (
                    ia_config.get("pdf_ingest_finished_at") if ia_config else None
                )
                ingest_error_detail = (
                    ia_config.get("pdf_ingest_error") if ia_config else None
                )

                if new_pdf:
                    ingest_state = "running"
                    ingest_started_at = datetime.utcnow()
                    ingest_finished_at = None
                    ingest_error_detail = None

                if ia_config:
                    c.execute(
                        """
                        UPDATE ia_config
                           SET model_name = %s,
                               model_token = %s,
                               system_prompt = %s,
                               enabled = %s,
                               business_description = %s,
                               followup_message_1 = %s,
                               followup_message_2 = %s,
                               followup_message_3 = %s,
                               followup_interval_minutes = %s,
                               followup_media_url_1 = %s,
                               followup_media_url_2 = %s,
                               followup_media_url_3 = %s,
                               followup_media_tipo_1 = %s,
                               followup_media_tipo_2 = %s,
                               followup_media_tipo_3 = %s,
                               pdf_filename = %s,
                               pdf_original_name = %s,
                               pdf_mime = %s,
                               pdf_size = %s,
                               pdf_uploaded_at = %s,
                               pdf_source_url = %s,
                               pdf_ingest_state = %s,
                               pdf_ingest_started_at = %s,
                               pdf_ingest_finished_at = %s,
                               pdf_ingest_error = %s
                         WHERE id = %s
                        """,
                        (
                            ia_model,
                            ia_token,
                            system_prompt,
                            ia_enabled,
                            business_description,
                            followup_message_1,
                            followup_message_2,
                            followup_message_3,
                            followup_interval_minutes,
                            followup_media_url_1,
                            followup_media_url_2,
                            followup_media_url_3,
                            followup_media_tipo_1,
                            followup_media_tipo_2,
                            followup_media_tipo_3,
                            new_pdf['stored_name'] if new_pdf else ia_config.get('pdf_filename'),
                            new_pdf['original_name'] if new_pdf else ia_config.get('pdf_original_name'),
                            new_pdf['mime'] if new_pdf else ia_config.get('pdf_mime'),
                            new_pdf['size'] if new_pdf else ia_config.get('pdf_size'),
                            datetime.utcnow() if new_pdf else ia_config.get('pdf_uploaded_at'),
                            new_pdf['source_url'] if new_pdf else ia_config.get('pdf_source_url'),
                            ingest_state,
                            ingest_started_at,
                            ingest_finished_at,
                            ingest_error_detail,
                            ia_config['id'],
                        ),
                    )
                else:
                    c.execute(
                        """
                        INSERT INTO ia_config
                            (model_name, model_token, system_prompt, enabled, business_description,
                             followup_message_1, followup_message_2, followup_message_3,
                             followup_interval_minutes,
                             followup_media_url_1, followup_media_url_2, followup_media_url_3,
                             followup_media_tipo_1, followup_media_tipo_2, followup_media_tipo_3,
                             pdf_filename, pdf_original_name,
                             pdf_mime, pdf_size, pdf_uploaded_at, pdf_source_url, pdf_ingest_state,
                             pdf_ingest_started_at, pdf_ingest_finished_at, pdf_ingest_error)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            ia_model,
                            ia_token,
                            system_prompt,
                            ia_enabled,
                            business_description,
                            followup_message_1,
                            followup_message_2,
                            followup_message_3,
                            followup_interval_minutes,
                            followup_media_url_1,
                            followup_media_url_2,
                            followup_media_url_3,
                            followup_media_tipo_1,
                            followup_media_tipo_2,
                            followup_media_tipo_3,
                            new_pdf['stored_name'] if new_pdf else None,
                            new_pdf['original_name'] if new_pdf else None,
                            new_pdf['mime'] if new_pdf else None,
                            new_pdf['size'] if new_pdf else None,
                            datetime.utcnow() if new_pdf else None,
                            new_pdf['source_url'] if new_pdf else None,
                            ingest_state,
                            ingest_started_at,
                            ingest_finished_at,
                            ingest_error_detail,
                        ),
                    )

                conn.commit()
                ia_config = _get_ia_config(c)
                pdf_url = None
                if ia_config and ia_config.get('pdf_filename'):
                    pdf_url = url_for(
                        'static',
                        filename=tenants.get_uploads_url_path(ia_config['pdf_filename'])
                    )
                if new_pdf:
                    status_message = (
                        "Catálogo guardado. Se está procesando en segundo plano."
                    )
                else:
                    status_message = 'Configuración de IA actualizada correctamente.'

                if (
                    new_pdf
                    and old_pdf_path
                    and os.path.exists(old_pdf_path)
                    and old_pdf_path != path
                ):
                    try:
                        os.remove(old_pdf_path)
                    except OSError:
                        pass

                if new_pdf and ia_config:
                    tenant = tenants.get_current_tenant()
                    if not enqueue_catalog_pdf_ingest(
                        config_id=ia_config["id"],
                        pdf_path=path,
                        stored_name=new_pdf["stored_name"],
                        tenant=tenant,
                    ):
                        error_message = (
                            "No se pudo iniciar el procesamiento del catálogo porque ya hay uno en curso."
                        )
                        status_message = None
                        c.execute(
                            """
                            UPDATE ia_config
                               SET pdf_ingest_state = %s,
                                   pdf_ingest_error = %s
                             WHERE id = %s
                            """,
                            ("failed", error_message, ia_config["id"]),
                        )
                        conn.commit()
                        ia_config = _get_ia_config(c)

        return render_template(
            'configuracion_ia.html',
            ia_config=ia_config,
            pdf_url=pdf_url,
            status_message=status_message,
            error_message=error_message,
        )
    finally:
        conn.close()


@config_bp.route('/configuracion/signup', methods=['GET'])
def configuracion_signup():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    oauth_code = (request.args.get("code") or "").strip()
    oauth_state = (request.args.get("state") or "").strip()
    if oauth_code:
        logger.info(
            "Código OAuth recibido en configuración de signup",
            extra={
                "tenant_key": tenants.get_active_tenant_key(),
                "oauth_code": oauth_code,
                "oauth_state": oauth_state,
            },
        )
        redirect_uri = _resolve_instagram_redirect_uri(request.base_url)
        result = _handle_instagram_oauth_code(oauth_code, redirect_uri)
        if not result.get("ok"):
            session["instagram_oauth_error"] = result.get("error") or "No se pudo procesar el OAuth de Instagram."
            if result.get("details"):
                session["instagram_oauth_error_details"] = result.get("details")
            logger.warning(
                "No se pudo procesar el código de Instagram OAuth",
                extra={"error": result.get("error"), "details": result.get("details")},
            )
        else:
            session["instagram_oauth_status"] = "Token de Instagram obtenido desde OAuth."
            session["instagram_oauth_token"] = result.get("access_token")
        return redirect(url_for("configuracion.configuracion_signup"))

    tenant = _resolve_signup_tenant()
    tenant_key = tenant.tenant_key if tenant else tenants.get_active_tenant_key()
    tenant_env = tenants.get_tenant_env(tenant)
    page_selection = _normalize_page_selection(tenant.metadata if tenant else None)
    messenger_page_selection = page_selection.get("messenger") or {}
    instagram_account = {}
    if tenant and isinstance(tenant.metadata, dict):
        instagram_account = tenant.metadata.get("instagram_account") or {}
    instagram_conversation_count, instagram_message_count = _fetch_instagram_backfill_counts(tenant)

    logger.info(
        "Renderizando signup embebido",
        extra={
            "tenant_key": tenant_key,
            "facebook_app_id_configured": bool(Config.FACEBOOK_APP_ID),
            "signup_config_code_present": bool(Config.WHATSAPP_EMBEDDED_SIGNUP_CONFIG_ID or Config.SIGNUP_FACEBOOK),
        },
    )
    signup_redirect_uri = _resolve_embedded_signup_redirect_uri(request.base_url)
    logger.info(
        "Redirect URI embebido resuelto",
        extra={
            "tenant_key": tenant_key,
            "redirect_uri": signup_redirect_uri,
        },
    )

    instagram_oauth_status = session.pop("instagram_oauth_status", "")
    instagram_oauth_token = session.pop("instagram_oauth_token", "")
    instagram_oauth_error = session.pop("instagram_oauth_error", "")
    instagram_oauth_error_details = session.pop("instagram_oauth_error_details", None)

    return render_template(
        'configuracion_signup.html',
        signup_config_code=(Config.WHATSAPP_EMBEDDED_SIGNUP_CONFIG_ID or Config.SIGNUP_FACEBOOK),
        messenger_embedded_code=Config.MESSENGER_EMBEDDED,
        facebook_app_id=Config.FACEBOOK_APP_ID,
        facebook_graph_api_version=Config.FACEBOOK_GRAPH_API_VERSION,
        signup_instagram_url=Config.SIGNUP_INSTRAGRAM,
        signup_redirect_uri=signup_redirect_uri,
        tenant_key=tenant_key,
        tenant_waba_id=tenant_env.get("WABA_ID"),
        tenant_phone_number_id=tenant_env.get("PHONE_NUMBER_ID"),
        messenger_page_id=_resolve_page_env_value("messenger", tenant_env),
        messenger_page_name=messenger_page_selection.get("page_name"),
        instagram_account_name=instagram_account.get("username") or instagram_account.get("id"),
        instagram_token_present=bool((tenant_env.get("INSTAGRAM_TOKEN") or "").strip()),
        instagram_conversation_count=instagram_conversation_count,
        instagram_message_count=instagram_message_count,
        instagram_oauth_status=instagram_oauth_status,
        instagram_oauth_token=instagram_oauth_token,
        instagram_oauth_error=instagram_oauth_error,
        instagram_oauth_error_details=instagram_oauth_error_details,
    )


@config_bp.route('/configuracion/instagram/callback', methods=['GET'])
def instagram_oauth_callback():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    oauth_code = (request.args.get("code") or "").strip()
    if not oauth_code:
        return redirect(url_for("configuracion.configuracion_signup"))

    redirect_uri = _resolve_instagram_redirect_uri(request.base_url)
    result = _handle_instagram_oauth_code(oauth_code, redirect_uri)
    if not result.get("ok"):
        session["instagram_oauth_error"] = result.get("error") or "No se pudo completar el OAuth de Instagram."
        if result.get("details"):
            session["instagram_oauth_error_details"] = result.get("details")
        logger.warning(
            "No se pudo completar el callback de Instagram OAuth",
            extra={"error": result.get("error"), "details": result.get("details")},
        )
    else:
        session["instagram_oauth_status"] = "Token de Instagram obtenido desde OAuth."
        session["instagram_oauth_token"] = result.get("access_token")
    return redirect(url_for("configuracion.configuracion_signup"))


@config_bp.route('/configuracion/instagram/reset', methods=['POST'])
def instagram_reset_signup():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["INSTAGRAM_TOKEN"] = None
    env_updates["INSTAGRAM_ACCOUNT_ID"] = None
    env_updates["INSTAGRAM_PAGE_ID"] = None
    tenants.update_tenant_env(tenant.tenant_key, env_updates)

    metadata_updates = {}
    raw_metadata = tenant.metadata if isinstance(tenant.metadata, dict) else {}
    if "instagram_account" in raw_metadata:
        metadata_updates["instagram_account"] = None

    page_selection = raw_metadata.get("page_selection")
    if isinstance(page_selection, dict):
        updated_selection = dict(page_selection)
        if "instagram" in updated_selection:
            updated_selection.pop("instagram", None)
        if (updated_selection.get("platform") or "").strip().lower() == "instagram":
            updated_selection.pop("platform", None)
            updated_selection.pop("page_id", None)
            updated_selection.pop("page_name", None)
        metadata_updates["page_selection"] = updated_selection or None
    elif page_selection is not None:
        metadata_updates["page_selection"] = None

    if metadata_updates:
        tenants.update_tenant_metadata(tenant.tenant_key, metadata_updates)

    session.pop("instagram_oauth_status", None)
    session.pop("instagram_oauth_token", None)
    session.pop("instagram_oauth_error", None)
    session.pop("instagram_oauth_error_details", None)

    logger.info(
        "Datos anteriores de Instagram limpiados antes de iniciar embedded signup",
        extra={"tenant_key": tenant.tenant_key},
    )
    return {"ok": True, "message": "Datos de Instagram limpiados."}


@config_bp.route('/configuracion/messenger/reset', methods=['POST'])
def messenger_reset_signup():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["MESSENGER_TOKEN"] = None
    env_updates["MESSENGER_PAGE_ID"] = None
    env_updates["MESSENGER_PAGE_ACCESS_TOKEN"] = None
    env_updates["PAGE_ID"] = None
    env_updates["PAGE_ACCESS_TOKEN"] = None
    if (env_updates.get("PLATFORM") or "").strip().lower() == "messenger":
        env_updates["PLATFORM"] = None
    tenants.update_tenant_env(tenant.tenant_key, env_updates)

    metadata_updates = {}
    raw_metadata = tenant.metadata if isinstance(tenant.metadata, dict) else {}
    if "messenger_embedded_signup" in raw_metadata:
        metadata_updates["messenger_embedded_signup"] = None

    page_selection = raw_metadata.get("page_selection")
    if isinstance(page_selection, dict):
        updated_selection = dict(page_selection)
        if "messenger" in updated_selection:
            updated_selection.pop("messenger", None)
        if (updated_selection.get("platform") or "").strip().lower() == "messenger":
            updated_selection.pop("platform", None)
            updated_selection.pop("page_id", None)
            updated_selection.pop("page_name", None)
        metadata_updates["page_selection"] = updated_selection or None
    elif page_selection is not None:
        metadata_updates["page_selection"] = None

    if metadata_updates:
        tenants.update_tenant_metadata(tenant.tenant_key, metadata_updates)

    logger.info(
        "Datos anteriores de Messenger limpiados desde integraciones",
        extra={"tenant_key": tenant.tenant_key},
    )
    return {"ok": True, "message": "Datos de Messenger limpiados."}


@config_bp.route('/configuracion/whatsapp/reset', methods=['POST'])
def whatsapp_reset_signup():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["META_TOKEN"] = None
    env_updates["LONG_LIVED_TOKEN"] = None
    env_updates["PHONE_NUMBER_ID"] = None
    env_updates["WABA_ID"] = None
    env_updates["BUSINESS_ID"] = None
    tenants.update_tenant_env(tenant.tenant_key, env_updates)

    metadata_updates = {}
    raw_metadata = tenant.metadata if isinstance(tenant.metadata, dict) else {}
    if "embedded_signup_payload" in raw_metadata:
        metadata_updates["embedded_signup_payload"] = None
    if "whatsapp_business" in raw_metadata:
        metadata_updates["whatsapp_business"] = None

    if metadata_updates:
        tenants.update_tenant_metadata(tenant.tenant_key, metadata_updates)

    logger.info(
        "Datos anteriores de WhatsApp limpiados desde integraciones",
        extra={"tenant_key": tenant.tenant_key},
    )
    return {"ok": True, "message": "Datos de WhatsApp limpiados."}


@config_bp.route('/configuracion/signup', methods=['POST'])
def save_signup():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        logger.warning("Signup embebido falló: tenant actual no encontrado")
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        logger.exception("Signup embebido falló al parsear el payload JSON")
        return {"ok": False, "error": "Payload inválido"}, 400

    logger.info(
        "Procesando signup embebido",
        extra={
            "tenant_key": tenant.tenant_key,
            "payload_keys": sorted(list(payload.keys())),
        },
    )
    embedded_code = (payload.get("code") or "").strip()
    provided_redirect_uri = (payload.get("redirect_uri") or "").strip()
    code_exchange_failed = False
    code_exchange_error = None
    if embedded_code:
        logger.info(
            "Código embebido recibido",
            extra={"tenant_key": tenant.tenant_key, "code": embedded_code},
        )
        resolved_redirect_uri = _resolve_embedded_signup_redirect_uri(request.base_url)
        redirect_uri = provided_redirect_uri or resolved_redirect_uri
        fallback_redirect_uri = None
        if provided_redirect_uri and resolved_redirect_uri and provided_redirect_uri != resolved_redirect_uri:
            fallback_redirect_uri = resolved_redirect_uri
        logger.info(
            "Redirect URI embebido para intercambio de token",
            extra={
                "tenant_key": tenant.tenant_key,
                "redirect_uri": redirect_uri,
                "fallback_redirect_uri": fallback_redirect_uri,
            },
        )
        token_response = _exchange_embedded_signup_code_with_fallbacks(
            embedded_code,
            redirect_uri,
            fallback_uri=fallback_redirect_uri,
        )
        if token_response.get("ok"):
            payload["access_token"] = token_response.get("access_token")
            logger.info(
                "Token embebido obtenido desde código",
                extra={
                    "tenant_key": tenant.tenant_key,
                    "access_token": token_response.get("access_token"),
                    "response": token_response.get("raw"),
                },
            )
        else:
            code_exchange_failed = True
            code_exchange_error = token_response.get("error") or "No se pudo intercambiar el código embebido."
            logger.warning(
                "No se pudo obtener el token desde el código embebido",
                extra={
                    "tenant_key": tenant.tenant_key,
                    "error": token_response.get("error"),
                    "details": token_response.get("details"),
                },
            )

    resolved_token = (payload.get("access_token") or payload.get("token") or "").strip()
    if code_exchange_failed and not resolved_token:
        details = token_response.get("details") if isinstance(token_response, dict) else None
        return {
            "ok": False,
            "error": _build_embedded_signup_error_message(code_exchange_error, details),
            "details": details,
        }, 400
    logger.info(
        "Token embebido recibido",
        extra={
            "tenant_key": tenant.tenant_key,
            "access_token": payload.get("access_token") or payload.get("token"),
        },
    )
    logger.info(
        "Payload completo de signup embebido",
        extra={
            "tenant_key": tenant.tenant_key,
            "payload": payload,
        },
    )

    current_env = tenants.get_tenant_env(tenant)
    env_updates = {key: current_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates.update(
        {
            "META_TOKEN": resolved_token or current_env.get("META_TOKEN"),
            "LONG_LIVED_TOKEN": (payload.get("access_token") or payload.get("long_lived_token") or "").strip()
            or current_env.get("LONG_LIVED_TOKEN"),
            "PHONE_NUMBER_ID": (payload.get("phone_number_id") or payload.get("phone_id") or "").strip()
            or current_env.get("PHONE_NUMBER_ID"),
            "WABA_ID": (payload.get("waba_id") or "").strip() or current_env.get("WABA_ID"),
            "BUSINESS_ID": (payload.get("business_id") or payload.get("business_manager_id") or "").strip()
            or current_env.get("BUSINESS_ID"),
        }
    )

    logger.info(
        "Actualizando entorno con datos del signup embebido",
        extra={
            "tenant_key": tenant.tenant_key,
            "has_meta_token": bool(env_updates.get("META_TOKEN")),
            "has_long_lived": bool(env_updates.get("LONG_LIVED_TOKEN")),
            "has_phone_number_id": bool(env_updates.get("PHONE_NUMBER_ID")),
            "has_waba_id": bool(env_updates.get("WABA_ID")),
            "has_business_id": bool(env_updates.get("BUSINESS_ID")),
        },
    )

    business_info = payload.get("business") or payload.get("business_info")
    metadata_updates = {
        "embedded_signup_payload": payload,
    }
    if isinstance(business_info, dict) and business_info:
        metadata_updates["whatsapp_business"] = business_info

    if metadata_updates:
        logger.info(
            "Guardando metadata de negocio desde signup embebido",
            extra={
                "tenant_key": tenant.tenant_key,
                "metadata_fields": sorted(list(metadata_updates.keys())),
            },
        )
    else:
        logger.info(
            "No se encontró metadata de negocio en el payload de signup",
            extra={"tenant_key": tenant.tenant_key},
        )

    tenants.update_tenant_env(tenant.tenant_key, env_updates)
    if metadata_updates:
        tenants.update_tenant_metadata(tenant.tenant_key, metadata_updates)

    return {
        "ok": True,
        "message": "Credenciales de WhatsApp actualizadas.",
        "env": tenants.get_tenant_env(tenant),
    }


@config_bp.route('/configuracion/whatsapp/phone-numbers', methods=['GET'])
def whatsapp_phone_numbers():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    token = tenant_env.get("META_TOKEN")
    waba_id = tenant_env.get("WABA_ID")

    response = list_phone_numbers(token, waba_id)
    status = 200 if response.get("ok") else 400
    return response, status


@config_bp.route('/configuracion/whatsapp/accounts', methods=['GET'])
def whatsapp_accounts():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    token, business_id, _tenant_env = _resolve_whatsapp_token_and_business_id(tenant)
    if not business_id:
        return {"ok": False, "error": "Falta BUSINESS_ID para consultar cuentas de WhatsApp."}, 400

    fields = "id,name,currency,owner_business_info"
    limit = request.args.get("limit", "20")

    client_accounts = _graph_get(
        f"{business_id}/client_whatsapp_business_accounts",
        token,
        params={"fields": fields, "limit": limit},
    )
    if not client_accounts.get("ok"):
        return client_accounts, 400

    owned_accounts = _graph_get(
        f"{business_id}/owned_whatsapp_business_accounts",
        token,
        params={"fields": fields, "limit": limit},
    )
    if not owned_accounts.get("ok"):
        return owned_accounts, 400

    return {
        "ok": True,
        "business_id": business_id,
        "client_accounts": _extract_graph_list(client_accounts.get("data") or {}),
        "owned_accounts": _extract_graph_list(owned_accounts.get("data") or {}),
    }


@config_bp.route('/configuracion/whatsapp/account-details', methods=['GET'])
def whatsapp_account_details():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    token, _business_id, tenant_env = _resolve_whatsapp_token_and_business_id(tenant)
    waba_id = (request.args.get("waba_id") or tenant_env.get("WABA_ID") or "").strip()
    if not waba_id:
        return {"ok": False, "error": "Falta WABA_ID para consultar el detalle de la cuenta."}, 400

    fields = request.args.get("fields") or "id,name,currency,owner_business_info"
    detail = _graph_get(waba_id, token, params={"fields": fields})
    if not detail.get("ok"):
        return detail, 400

    return {"ok": True, "waba_id": waba_id, "account": detail.get("data")}


@config_bp.route('/configuracion/whatsapp/message-templates', methods=['GET'])
def whatsapp_message_templates():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    token, _business_id, tenant_env = _resolve_whatsapp_token_and_business_id(tenant)
    waba_id = (request.args.get("waba_id") or tenant_env.get("WABA_ID") or "").strip()
    if not waba_id:
        return {"ok": False, "error": "Falta WABA_ID para consultar plantillas."}, 400

    fields = request.args.get("fields") or "language,name,rejected_reason,status,category,sub_category,last_updated_time,components,quality_score"
    limit = request.args.get("limit", "50")
    templates = _graph_get(
        f"{waba_id}/message_templates",
        token,
        params={"fields": fields, "limit": limit},
    )
    if not templates.get("ok"):
        return templates, 400

    return {"ok": True, "waba_id": waba_id, "templates": _extract_graph_list(templates.get("data") or {})}


@config_bp.route('/configuracion/whatsapp/subscribed-apps', methods=['GET'])
def whatsapp_subscribed_apps():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    token, _business_id, tenant_env = _resolve_whatsapp_token_and_business_id(tenant)
    waba_id = (request.args.get("waba_id") or tenant_env.get("WABA_ID") or "").strip()
    if not waba_id:
        return {"ok": False, "error": "Falta WABA_ID para consultar apps suscritas."}, 400

    response = _graph_get(f"{waba_id}/subscribed_apps", token)
    if not response.get("ok"):
        return response, 400

    return {"ok": True, "waba_id": waba_id, "apps": _extract_graph_list(response.get("data") or {})}


@config_bp.route('/configuracion/whatsapp/subscribe-app', methods=['POST'])
def whatsapp_subscribe_app():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        payload = {}

    token, _business_id, tenant_env = _resolve_whatsapp_token_and_business_id(tenant)
    waba_id = (payload.get("waba_id") or tenant_env.get("WABA_ID") or "").strip()
    if not waba_id:
        return {"ok": False, "error": "Falta WABA_ID para suscribir la app."}, 400

    response = _graph_post(f"{waba_id}/subscribed_apps", token)
    if not response.get("ok"):
        return response, 400

    return {"ok": True, "waba_id": waba_id, "result": response.get("data")}


@config_bp.route('/configuracion/whatsapp/phone-number', methods=['POST'])
def whatsapp_save_phone_number():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    phone_number_id = (payload.get("phone_number_id") or "").strip()
    if not phone_number_id:
        return {"ok": False, "error": "Selecciona un número válido."}, 400

    current_env = tenants.get_tenant_env(tenant)
    env_updates = {key: current_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["PHONE_NUMBER_ID"] = phone_number_id

    tenants.update_tenant_env(tenant.tenant_key, env_updates)

    return {
        "ok": True,
        "message": "Número de WhatsApp actualizado.",
        "env": tenants.get_tenant_env(tenant),
    }


@config_bp.route('/configuracion/whatsapp/phone-number-action', methods=['POST'])
def whatsapp_phone_number_action():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    current_env = tenants.get_tenant_env(tenant)
    token = (current_env.get("META_TOKEN") or current_env.get("LONG_LIVED_TOKEN") or "").strip()
    phone_number_id = (payload.get("phone_number_id") or current_env.get("PHONE_NUMBER_ID") or "").strip()
    action = (payload.get("action") or "").strip().lower()

    graph_response = _whatsapp_phone_number_action(phone_number_id, action, token)
    if not graph_response.get("ok"):
        return graph_response, 400

    action_label = "registrado" if action == "register" else "eliminado"
    return {
        "ok": True,
        "action": action,
        "phone_number_id": phone_number_id,
        "message": f"Número {action_label} correctamente.",
        "result": graph_response.get("data"),
    }


@config_bp.route('/configuracion/messenger/pages', methods=['POST'])
def messenger_pages():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    platform = (payload.get("platform") or "").strip().lower() or "messenger"
    if platform != "messenger":
        return {"ok": False, "error": "Solo puedes consultar páginas de Messenger."}, 400
    provided_token = (payload.get("user_access_token") or "").strip()
    provided_page_token = (payload.get("page_access_token") or "").strip()
    tenant_env = tenants.get_tenant_env(tenant)
    token = _resolve_page_user_token(platform, tenant_env, provided_token)

    if provided_token:
        env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
        env_updates[_resolve_user_token_key(platform)] = provided_token
        tenants.update_tenant_env(tenant.tenant_key, env_updates)

    if provided_page_token:
        response = _fetch_page_from_token(provided_page_token)
        if not response.get("ok"):
            return response, 400
        page = response.get("page") or {}
        return {
            "ok": True,
            "pages": [{"id": page.get("id"), "name": page.get("name")}],
        }

    response = _fetch_page_accounts(token)
    if not response.get("ok"):
        return response, 400

    pages = response.get("pages", [])
    return {"ok": True, "pages": [{"id": page["id"], "name": page.get("name")} for page in pages]}


@config_bp.route('/configuracion/messenger/signup', methods=['POST'])
def messenger_signup():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    embedded_code = (payload.get("code") or "").strip()
    access_token = (payload.get("access_token") or payload.get("token") or "").strip()
    provided_redirect_uri = (payload.get("redirect_uri") or "").strip()
    if embedded_code:
        resolved_redirect_uri = _resolve_embedded_signup_redirect_uri(
            url_for("configuracion.configuracion_signup", _external=True)
        )
        redirect_uri = provided_redirect_uri or resolved_redirect_uri
        fallback_redirect_uri = None
        if provided_redirect_uri and resolved_redirect_uri and provided_redirect_uri != resolved_redirect_uri:
            fallback_redirect_uri = resolved_redirect_uri

        token_response = _exchange_embedded_signup_code_with_fallbacks(
            embedded_code,
            redirect_uri,
            fallback_uri=fallback_redirect_uri,
        )
        if token_response.get("ok"):
            access_token = token_response.get("access_token") or access_token
            payload["token_exchange"] = token_response.get("raw")
        else:
            details = token_response.get("details") or {}
            error_detail = ""
            if isinstance(details, dict):
                meta_error = details.get("error")
                if isinstance(meta_error, dict):
                    error_detail = (
                        meta_error.get("error_user_msg")
                        or meta_error.get("message")
                        or ""
                    )
            error_message = _build_embedded_signup_error_message(
                token_response.get("error") or "No se pudo intercambiar el código de Messenger.",
                token_response.get("details"),
            )
            if error_detail and error_detail not in error_message:
                error_message = f"{error_message} {error_detail}".strip()
            return {
                "ok": False,
                "error": error_message,
                "details": token_response.get("details"),
            }, 400

    if not access_token:
        return {"ok": False, "error": "No se obtuvo un token de Messenger."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["MESSENGER_TOKEN"] = access_token
    env_updates["PLATFORM"] = "messenger"
    tenants.update_tenant_env(tenant.tenant_key, env_updates)
    tenants.update_tenant_metadata(
        tenant.tenant_key,
        {"messenger_embedded_signup": payload},
    )

    logger.info(
        "Token de Messenger actualizado desde Embedded Signup",
        extra={
            "tenant_key": tenant.tenant_key,
            "has_token": bool(access_token),
        },
    )

    return {
        "ok": True,
        "message": "Token de Messenger actualizado.",
        "has_token": bool(access_token),
    }


@config_bp.route('/configuracion/messenger/page', methods=['POST'])
def messenger_save_page():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    page_id = (payload.get("page_id") or "").strip()
    platform = (payload.get("platform") or "").strip().lower()
    provided_token = (payload.get("user_access_token") or "").strip()
    provided_page_token = (payload.get("page_access_token") or "").strip()

    if platform != "messenger":
        return {"ok": False, "error": "Selecciona una plataforma válida."}, 400

    tenant_env = tenants.get_tenant_env(tenant)
    page_entry = None
    if provided_page_token:
        response = _fetch_page_from_token(provided_page_token)
        if not response.get("ok"):
            return response, 400
        page_entry = response.get("page")
        if not page_entry:
            return {"ok": False, "error": "No se pudo obtener la página del token."}, 400
        if page_id and str(page_entry.get("id")) != page_id:
            return {"ok": False, "error": "El token no corresponde a la página indicada."}, 400
    else:
        if not page_id:
            return {"ok": False, "error": "Selecciona una página válida."}, 400
        token = _resolve_page_user_token(platform, tenant_env, provided_token)
        response = _fetch_page_accounts(token)
        if not response.get("ok"):
            return response, 400

        for page in response.get("pages", []):
            if str(page.get("id")) == page_id:
                page_entry = page
                break

        if not page_entry or not page_entry.get("access_token"):
            return {"ok": False, "error": "No se pudo obtener el token de la página."}, 400

    if not page_entry or not page_entry.get("access_token"):
        return {"ok": False, "error": "No se pudo obtener el token de la página."}, 400

    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates[_resolve_page_env_key(platform, "PAGE_ID")] = page_entry.get("id")
    env_updates[_resolve_page_env_key(platform, "PAGE_ACCESS_TOKEN")] = page_entry.get("access_token")
    env_updates["PLATFORM"] = platform
    if provided_token:
        env_updates[_resolve_user_token_key(platform)] = provided_token

    tenants.update_tenant_env(tenant.tenant_key, env_updates)

    page_selection = _normalize_page_selection(tenant.metadata if tenant else None)
    page_selection[platform] = {
        "page_id": page_entry.get("id"),
        "page_name": page_entry.get("name"),
    }
    tenants.update_tenant_metadata(
        tenant.tenant_key,
        {"page_selection": page_selection},
    )

    return {
        "ok": True,
        "message": "Página actualizada correctamente.",
        "page": {
            "id": page_entry.get("id"),
            "name": page_entry.get("name"),
            "platform": platform,
        },
    }


@config_bp.route('/configuracion/instagram/token', methods=['POST'])
def instagram_save_token():
    if not _require_admin():
        return {"ok": False, "error": "No autorizado"}, 403

    tenant = _resolve_signup_tenant()
    if not tenant:
        return {"ok": False, "error": "No se encontró la empresa actual."}, 400

    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return {"ok": False, "error": "Payload inválido"}, 400

    user_token = (payload.get("user_access_token") or "").strip()
    if not user_token:
        return {"ok": False, "error": "Ingresa un token de Instagram válido."}, 400

    response = _fetch_instagram_user(user_token)
    if not response.get("ok"):
        return response, 400

    account = response.get("account") or {}
    tenant_env = tenants.get_tenant_env(tenant)
    env_updates = {key: tenant_env.get(key) for key in tenants.TENANT_ENV_KEYS}
    env_updates["INSTAGRAM_TOKEN"] = user_token
    if account.get("id"):
        env_updates["INSTAGRAM_ACCOUNT_ID"] = account.get("id")
    if account.get("user_id") or account.get("id"):
        env_updates["INSTAGRAM_PAGE_ID"] = account.get("user_id") or account.get("id")
    tenants.update_tenant_env(tenant.tenant_key, env_updates)
    tenants.trigger_page_backfill_for_platform(tenant, "instagram")
    tenants.update_tenant_metadata(
        tenant.tenant_key,
        {"instagram_account": account},
    )
    logger.info(
        "Token de Instagram actualizado",
        extra={
            "tenant_key": tenant.tenant_key,
            "instagram_account_id": account.get("id"),
            "instagram_username": account.get("username"),
        },
    )

    return {
        "ok": True,
        "message": "Token de Instagram actualizado.",
        "account": account,
    }

@config_bp.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    return _reglas_view('configuracion.html')

@config_bp.route('/reglas', methods=['GET', 'POST'])
def reglas():
    return _reglas_view('reglas.html')

@config_bp.route('/eliminar_regla/<int:regla_id>', methods=['POST'])
def eliminar_regla(regla_id):
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM reglas WHERE id = %s", (regla_id,))
        conn.commit()
        return redirect(url_for('configuracion.reglas'))
    finally:
        conn.close()

@config_bp.route('/botones', methods=['GET', 'POST'])
def botones():
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        opciones_expr = _botones_opciones_column(c, conn)
        categoria_expr = _botones_categoria_column(c, conn)
        if request.method == 'POST':
            selected_users = _selected_user_ids(request.form, c)
            # Importar botones desde Excel
            if 'archivo' in request.files and request.files['archivo']:
                archivo = request.files['archivo']
                wb = load_workbook(archivo)
                hoja = wb.active
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    if not fila:
                        continue
                    nombre = fila[0]
                    mensaje = fila[1] if len(fila) > 1 else None
                    tipo = fila[2] if len(fila) > 2 else None
                    media_url = fila[3] if len(fila) > 3 else None
                    opciones = fila[4] if len(fila) > 4 else None
                    categoria = fila[5] if len(fila) > 5 else None
                    if isinstance(opciones, (dict, list)):
                        opciones = json.dumps(opciones, ensure_ascii=False)
                    elif opciones is not None:
                        opciones = str(opciones).strip()
                        if not opciones:
                            opciones = None
                    medias = []
                    if media_url:
                        urls = [u.strip() for u in re.split(r'[\n,]+', str(media_url)) if u and u.strip()]
                        for url in urls:
                            ok, mime = _url_ok(url)
                            if ok:
                                medias.append((url, mime))
                    if mensaje:
                        c.execute(
                            "INSERT INTO botones (nombre, mensaje, tipo, opciones, categoria) VALUES (%s, %s, %s, %s, %s)",
                            (nombre, mensaje, tipo, opciones, categoria)
                        )
                        boton_id = c.lastrowid
                        for url, mime in medias:
                            c.execute(
                                "INSERT INTO boton_medias (boton_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                                (boton_id, url, mime)
                            )
                        _assign_boton_users(c, boton_id, selected_users)
                conn.commit()
            elif request.form.get('regla_id'):
                regla_id = request.form.get('regla_id')
                try:
                    regla_id = int(regla_id)
                except (TypeError, ValueError):
                    regla_id = None

                if regla_id:
                    c.execute(
                        """
                        SELECT r.respuesta,
                               r.tipo,
                               r.opciones,
                               GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                               GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
                          FROM reglas r
                          LEFT JOIN regla_medias m ON r.id = m.regla_id
                         WHERE r.id = %s
                         GROUP BY r.id
                        """,
                        (regla_id,)
                    )
                    row = c.fetchone()
                else:
                    row = None

                if row and row[0]:
                    nombre = request.form.get('nombre')
                    respuesta = row[0]
                    tipo = row[1] or 'texto'
                    opciones_raw = row[2]
                    media_urls_raw = row[3].split('||') if row[3] else []
                    media_tipos_raw = row[4].split('||') if row[4] else []
                    medias = []
                    for idx, url in enumerate(media_urls_raw):
                        if not url:
                            continue
                        mime = media_tipos_raw[idx] if idx < len(media_tipos_raw) else None
                        medias.append((url, mime))

                    opciones_value = opciones_raw if opciones_raw else None
                    c.execute(
                        "INSERT INTO botones (nombre, mensaje, tipo, opciones, categoria) VALUES (%s, %s, %s, %s, %s)",
                        (nombre, respuesta, tipo, opciones_value, request.form.get('categoria'))
                    )
                    boton_id = c.lastrowid
                    for url, mime in medias:
                        c.execute(
                            "INSERT INTO boton_medias (boton_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                            (boton_id, url, mime)
                        )
                    _assign_boton_users(c, boton_id, selected_users)
                    conn.commit()
            # Agregar botón manual
            elif 'mensaje' in request.form:
                nombre = request.form.get('nombre')
                nuevo_mensaje = request.form['mensaje']
                tipo = request.form.get('tipo')
                media_files = request.files.getlist('media')
                medias = []
                for media_file in media_files:
                    if media_file and media_file.filename:
                        filename = secure_filename(media_file.filename)
                        unique = f"{uuid.uuid4().hex}_{filename}"
                        path = os.path.join(_media_root(), unique)
                        media_file.save(path)
                        url = url_for(
                            'static',
                            filename=tenants.get_uploads_url_path(unique),
                            _external=True,
                        )
                        medias.append((url, media_file.mimetype.split(';', 1)[0]))
                media_url = request.form.get('media_url', '')
                urls = [u.strip() for u in re.split(r'[\n,]+', media_url) if u and u.strip()]
                for url in urls:
                    ok, mime = _url_ok(url)
                    if ok:
                        medias.append((url, mime))
                if nuevo_mensaje:
                    c.execute(
                        "INSERT INTO botones (nombre, mensaje, tipo, opciones, categoria) VALUES (%s, %s, %s, %s, %s)",
                        (nombre, nuevo_mensaje, tipo, None, request.form.get('categoria'))
                    )
                    boton_id = c.lastrowid
                    for url, mime in medias:
                        c.execute(
                            "INSERT INTO boton_medias (boton_id, media_url, media_tipo) VALUES (%s, %s, %s)",
                            (boton_id, url, mime)
                        )
                    _assign_boton_users(c, boton_id, selected_users)
                    conn.commit()

        c.execute(
            f"""
            SELECT b.id, b.mensaje, b.tipo, b.nombre, {opciones_expr}, {categoria_expr},
                   GROUP_CONCAT(DISTINCT u.id ORDER BY u.username SEPARATOR '||') AS user_ids,
                   GROUP_CONCAT(DISTINCT u.username ORDER BY u.username SEPARATOR '||') AS usernames,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
              FROM botones b
              LEFT JOIN boton_usuarios bu ON b.id = bu.boton_id
              LEFT JOIN usuarios u ON bu.user_id = u.id
              LEFT JOIN boton_medias m ON b.id = m.boton_id
             GROUP BY b.id
             ORDER BY b.id
            """
        )
        botones = []
        for row in c.fetchall():
            user_ids = [int(uid) for uid in row[6].split('||')] if row[6] else []
            usernames = row[7].split('||') if row[7] else []
            media_urls = row[8].split('||') if row[8] else []
            media_tipos = row[9].split('||') if row[9] else []
            if media_urls:
                items = []
                for idx, url in enumerate(media_urls):
                    mime = media_tipos[idx] if idx < len(media_tipos) else ''
                    texto = f"{url} ({mime})" if mime else url
                    items.append(f"<li>{texto}</li>")
                media_urls_display = f"<ul>{''.join(items)}</ul>"
            else:
                media_urls_display = ''
            botones.append({
                'id': row[0],
                'mensaje': row[1] or '',
                'tipo': row[2] or 'texto',
                'nombre': row[3],
                'opciones': row[4] or '',
                'categoria': row[5],
                'user_ids': user_ids,
                'usernames': usernames,
                'media_urls': media_urls,
                'media_tipos': media_tipos,
                'media_urls_display': media_urls_display,
            })
        c.execute(
            """
            SELECT r.id, r.step, r.input_text, r.respuesta, r.tipo,
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
              FROM reglas r
              LEFT JOIN regla_medias m ON r.id = m.regla_id
             GROUP BY r.id
             ORDER BY r.step, r.id
            """
        )
        reglas = []
        for row in c.fetchall():
            reglas.append({
                'id': row[0],
                'step': row[1] or '',
                'input_text': row[2] or '',
                'respuesta': row[3] or '',
                'tipo': row[4] or '',
                'media_urls': row[5] or '',
                'media_tipos': row[6] or '',
            })
        c.execute("SELECT id, username FROM usuarios ORDER BY username")
        usuarios = [{'id': row[0], 'username': row[1]} for row in c.fetchall()]
        return render_template('botones.html', botones=botones, reglas=reglas, usuarios=usuarios)
    finally:
        conn.close()

@config_bp.route('/eliminar_boton/<int:boton_id>', methods=['POST'])
def eliminar_boton(boton_id):
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM botones WHERE id = %s", (boton_id,))
        conn.commit()
        return redirect(url_for('configuracion.botones'))
    finally:
        conn.close()

@config_bp.route('/botones/<int:boton_id>/usuarios', methods=['POST'])
def actualizar_boton_usuarios(boton_id):
    if not _require_admin():
        return redirect(url_for("auth.login"))

    conn = get_connection()
    c = conn.cursor()
    try:
        user_ids = _selected_user_ids(request.form, c, default_to_session=False)
        c.execute("DELETE FROM boton_usuarios WHERE boton_id = %s", (boton_id,))
        if user_ids:
            _assign_boton_users(c, boton_id, user_ids)
        conn.commit()
        return redirect(url_for('configuracion.botones'))
    finally:
        conn.close()

@config_bp.route('/get_botones')
def get_botones():
    if "user" not in session:
        return jsonify([]), 401

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("SELECT id FROM usuarios WHERE username = %s", (session.get("user"),))
        user_row = c.fetchone()
        if not user_row:
            return jsonify([])
        user_id = user_row[0]
        opciones_expr = _botones_opciones_column(c, conn)
        categoria_expr = _botones_categoria_column(c, conn)
        c.execute(
            f"""
            SELECT b.id, b.mensaje, b.tipo, b.nombre, {opciones_expr}, {categoria_expr},
                   GROUP_CONCAT(m.media_url SEPARATOR '||') AS media_urls,
                   GROUP_CONCAT(m.media_tipo SEPARATOR '||') AS media_tipos
              FROM botones b
              INNER JOIN boton_usuarios bu ON b.id = bu.boton_id
              LEFT JOIN boton_medias m ON b.id = m.boton_id
             WHERE bu.user_id = %s
             GROUP BY b.id
             ORDER BY b.id
            """,
            (user_id,)
        )
        rows = c.fetchall()
        return jsonify([
            {
                'id': r[0],
                'mensaje': r[1] or '',
                'tipo': r[2] or 'texto',
                'nombre': r[3],
                'opciones': r[4] or '',
                'categoria': r[5],
                'media_urls': r[6].split('||') if r[6] else [],
                'media_tipos': r[7].split('||') if r[7] else []
            }
            for r in rows
        ])
    finally:
        conn.close()
